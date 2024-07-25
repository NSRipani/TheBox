# libreria para usar patrones, utilizados para buscar y manipular cadenas de texto de manera flexible
import re

# Librería para generar Archivos de tipo Excel(.xlsx)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, numbers

# Librería de MySQL
import mysql.connector 
from mysql.connector import Error

# Librerías de PyQt6
from PyQt6.QtWidgets import (QLabel, QFileDialog, QCompleter, QAbstractScrollArea, QHeaderView, QGridLayout, QHBoxLayout, QDateEdit, 
                             QMessageBox, QTableWidget, QAbstractItemView, QTableWidgetItem, QPushButton, QLineEdit, QStatusBar, QWidget,
                             QVBoxLayout, QGroupBox, QMainWindow, QFrame, QTabWidget,QComboBox)
from PyQt6.QtGui import QIcon, QKeySequence, QAction, QPixmap
from PyQt6.QtCore import *

# Módulo de para las cajas de mensajes
from modulos.mensajes import ingreso_datos,mensaje_ingreso_datos, errorConsulta, inicio

# Validaciones y demas funciones 
from validaciones.usuario import (registroUSER, actualizacionUSER, limpiasElementosUser, limpiar_campos, limpiasElementosUseraActualizar, 
                                  autoCompletadoACTULIZAR,limpiar_tablaRecord, limpiar_tablaUpdate, tabla_registroUSER)
from validaciones.updateYdelete_usuario import tabla_updateUSER, tabla_eliminarUSER, borrarTabla
from validaciones.archivo_Excel import (tabla_registroUSUARIO, tabla_registroDISCIPLINA, horas_Excel, 
                                        tabla_libroDiario_CONTABILIDAD, pagos_EXCEL, excelConsulta)
from validaciones.disciplina import guardarACTIVIDAD, completar_CAMPOS_ACTIVIDAD, clear_tabla_disciplina, tabla_DISCIPLINA
from validaciones.pagos import seleccionDeTablaPAGOS, tabla_pagos
from validaciones.contabilidad import validaciones, tabla_contabilidad, selccionarTabla, limpiarCampos, clear_tabla, autocompletoDEBE,autocompletoHABER
from validaciones.horas import tabla_HorasTotales,tabla_HorasXEmpleado, autoCompletado, clearTabla
from validaciones.consultas import consultaPorAlumno, totalAlumno, limpiar, consultarDeDisciplina, consultaPorDisciplina, asistenciaTotal,asistenciaPorAlumno

# Módulo de Registro
from modulos.asistencia import Asistencia
from modulos.carga_empleado.reg_empleado import Empleado
from modulos.cargar_cuenta.carga_cuenta import CuentaContable
from modulos.archivoTexto.archivoTexto import ArchivoTexto

# Módulo de Estilos
from qss import style

# conexion
from conexion_DB.dataBase import conectar_base_de_datos

class VentanaPrincipal(QMainWindow):
    def __init__(self):
        super().__init__()
        self.ventana_pricipal()
        self.show()
        self.input_nombre1.setFocus()
        self.eliminacion_pendiente = False  # Atributo para controlar el estado del botón
        
    def onChangeTab(self,i):
        #  Conexión a la base de datos MySQL
        conn = conectar_base_de_datos()
        cursor = conn.cursor()
        print(i)
        match i:
            case 0:
                self.input_nombre1.setFocus()
                
                # Consulta para obtener los datos de una columna específica
                cursor.execute("SELECT nombre FROM usuario WHERE habilitado = 1")
                datos = cursor.fetchall()
                sugerencia = [str(item[0]) for item in datos]

                lista_nombre = QCompleter(sugerencia)
                lista_nombre.setCaseSensitivity(Qt.CaseSensitivity.CaseInsensitive)
                lista_nombre.setCompletionMode(QCompleter.CompletionMode.PopupCompletion)
                lista_nombre.popup().setStyleSheet(style.completer)
                self.input_nombre1.setCompleter(lista_nombre)
                return
            case 1:
                self.input_nombre2.setFocus()
                
                # Consulta para obtener los datos de una columna específica
                cursor.execute("SELECT nombre FROM usuario WHERE habilitado = 1")
                datos = cursor.fetchall()
                suger = [str(item[0]) for item in datos]

                lista_nombres = QCompleter(suger)
                lista_nombres.setCaseSensitivity(Qt.CaseSensitivity.CaseInsensitive)
                lista_nombres.setCompletionMode(QCompleter.CompletionMode.PopupCompletion)
                lista_nombres.popup().setStyleSheet(style.completer)
                self.input_nombre2.setCompleter(lista_nombres)
                return
            case 2:
                self.nombre_buscar3.setFocus()

                # Consulta para obtener los datos de una columna específica
                cursor.execute("SELECT nombre FROM usuario WHERE habilitado = 1")
                datos = cursor.fetchall()
                suger2 = [str(item[0]) for item in datos]

                lista_nombre = QCompleter(suger2)
                lista_nombre.setCaseSensitivity(Qt.CaseSensitivity.CaseInsensitive)
                lista_nombre.setCompletionMode(QCompleter.CompletionMode.PopupCompletion)
                lista_nombre.popup().setStyleSheet(style.completer)
                self.nombre_buscar3.setCompleter(lista_nombre)
                return
            case 4:
                self.idUser.setFocus()
                
                # ---- Para 'DNI' ---
                # Consulta para obtener los datos de una columna específica
                cursor.execute("SELECT dni FROM usuario WHERE habilitado = 1 ORDER BY dni ASC")
                dato = cursor.fetchall()
                self.sugerenciass = [str(item[0]) for item in dato]
                
                # Almacenar sugerencias en un conjunto para verificación rápida
                self.sugerencias_set = set(self.sugerenciass)
                
                idDNI = QCompleter(self.sugerenciass)
                idDNI.setCaseSensitivity(Qt.CaseSensitivity.CaseInsensitive)
                idDNI.setCompletionMode(QCompleter.CompletionMode.PopupCompletion)
                idDNI.popup().setStyleSheet(style.completer)
                self.idUser.setCompleter(idDNI)
                
                # Para 'DISCIPLINA'
                cursor.execute("SELECT id_disciplina, nombre, precio FROM disciplina WHERE habilitado = 1 ORDER BY nombre ASC")
                dato = cursor.fetchall()
                self.disciplinas = {str(item[1]): (item[0], item[2]) for item in dato}  # Mapear nombre a (id_disciplina, precio)
                
                sugerencias = list(self.disciplinas.keys())
                
                # Almacenar sugerencias en un conjunto para verificación rápida
                self.sugeren_set = set(sugerencias)
                
                idDis = QCompleter(sugerencias)
                idDis.setCaseSensitivity(Qt.CaseSensitivity.CaseInsensitive)
                idDis.setCompletionMode(QCompleter.CompletionMode.PopupCompletion)
                idDis.popup().setStyleSheet(style.completer)
                self.idDis.setCompleter(idDis)
                
                # Muestra el precio de cada disciplina al elegitr la disciplina
                idDis.activated.connect(self.actualizar_precio)
                
                # Para 'MEDIO DE PAGO'
                medio_de_pago = ["Efectivo", "Transferencia"]
                completer2 = QCompleter(medio_de_pago)
                completer2.setCaseSensitivity(Qt.CaseSensitivity.CaseInsensitive)
                completer2.setCompletionMode(QCompleter.CompletionMode.PopupCompletion)
                completer2.popup().setStyleSheet(style.completer)
                self.input_tipoDePago.setCompleter(completer2)
                
                return
            case 5:
                self.id_horas_empleado.setFocus()
                
                # Definir el atributo id_empleado_seleccionado
                self.id_empleado_seleccionado = None
                
                # ---- Para 'EMPLEADO' ----
                # Consulta para obtener los datos de una columna específica
                cursor.execute("SELECT id_empleado, nombre FROM registro_empleado WHERE habilitado = 1 ORDER BY nombre ASC")
                datos = cursor.fetchall()
                self.listas = [(str(item[1]), item[0]) for item in datos]  # Crear una lista de tuplas (nombre, id_empleado)
                
                lista_empleado = QCompleter([item[0] for item in self.listas])  # Usar solo los nombres para la lista desplegable
                lista_empleado.setCaseSensitivity(Qt.CaseSensitivity.CaseInsensitive)
                lista_empleado.setCompletionMode(QCompleter.CompletionMode.PopupCompletion)
                lista_empleado.popup().setStyleSheet(style.completer)
                self.id_horas_empleado.setCompleter(lista_empleado)
                
                # Agregar evento de selección de elemento del completer
                lista_empleado.activated[str].connect(self.guardar_id_empleado)#[str]
                return
            case 6:
                self.view_nomb.setFocus()
                
                # --- Para 'BALANCE' --- Actualización para la USUARIO
                cursor.execute("SELECT u.dni FROM usuario u WHERE habilitado = 1 ORDER BY u.dni ASC")        
                datos = cursor.fetchall()
                self.numeros_dni = [str(item[0]) for item in datos]

                nombre = QCompleter(self.numeros_dni)
                nombre.setCaseSensitivity(Qt.CaseSensitivity.CaseInsensitive)
                nombre.setCompletionMode(QCompleter.CompletionMode.PopupCompletion)
                nombre.popup().setStyleSheet(style.completer)
                self.view_nomb.setCompleter(nombre)

                # --- Para 'BALANCE' --- Actualización para la DISCIPLINA
                cursor.execute("SELECT d.nombre FROM disciplina d WHERE habilitado = 1 ORDER BY d.nombre ASC")
                datos = cursor.fetchall()
                self.activi = [str(item[0]) for item in datos]

                actividad = QCompleter(self.activi)
                actividad.setCaseSensitivity(Qt.CaseSensitivity.CaseInsensitive)
                actividad.setCompletionMode(QCompleter.CompletionMode.PopupCompletion)
                actividad.popup().setStyleSheet(style.completer)
                self.view_disciplina.setCompleter(actividad)
                return 
            case 7:
                self.concepto_debe.setFocus()
                
                
                # --- Para CONTABILIDAD --- CUENTA DEBE
                # Consulta para obtener los datos de una columna específica
                cursor.execute("SELECT nombre FROM cuenta WHERE habilitado = 1 AND categoria = 'debe'")
                dato = cursor.fetchall()
                self.sugerencia1 = [str(item[0]) for item in dato]
                
                lista_debe = QCompleter(self.sugerencia1)
                lista_debe.setCaseSensitivity(Qt.CaseSensitivity.CaseInsensitive)
                lista_debe.setCompletionMode(QCompleter.CompletionMode.PopupCompletion)
                lista_debe.popup().setStyleSheet(style.completer)
                self.concepto_debe.setCompleter(lista_debe)
                
                # --- Para CONTABILIDAD --- CUENTA HABER
                # Consulta para obtener los datos de una columna específica
                cursor.execute("SELECT nombre FROM cuenta WHERE habilitado = 1 AND categoria = 'haber'")
                datos = cursor.fetchall()
                self.sugerencia2 = [str(item[0]) for item in datos]

                lista_haber = QCompleter(self.sugerencia2)
                lista_haber.setCaseSensitivity(Qt.CaseSensitivity.CaseInsensitive)
                lista_haber.setCompletionMode(QCompleter.CompletionMode.PopupCompletion)
                lista_haber.popup().setStyleSheet(style.completer)
                self.concepto_haber.setCompleter(lista_haber)
                                
                return
            
        cursor.close()
        conn.close()
    
    def ventana_pricipal(self):        
        self.setWindowIcon(QIcon("img/logo.png"))
        self.setWindowTitle("The Box - Gestion de usuarios")
        # self.setWindowState(Qt.WindowState.WindowMaximized)
        
        #BARRA INFERIOR DE ESTADO
        self.status_Bar = QStatusBar()
        self.status_Bar.setStyleSheet(style.estilo_statusbar)
        self.status_Bar.showMessage("by: Nicolas S. Ripani - 2024")
        self.status_Bar.showMaximized()
        self.setStatusBar(self.status_Bar)
        
        # FUNCION PARA EL "MENUBAR"
        self.acciones()
        
        #VARIABLE PARA CREAR UN CONTENEDOR
        layout_vertical1 = QVBoxLayout()
                
        # CREACION DE UN LOGOTIPO
        logo_label = QLabel(self)
        logo_label.setStyleSheet("background-color: transparent;")
        pixmap = QPixmap("img/logo.png")
        pixmap = pixmap.scaled(200, 200)
        logo_label.setPixmap(pixmap)
        
        # BOTONES QUE SERAN VINCULADOS CON EL TabWIDGET
        registrar_button = QPushButton(" REGISTRAR", self)
        registrar_button.setStyleSheet(style.estilo)
        registrar_button.setCursor(Qt.CursorShape.PointingHandCursor)
        registrar_button.setFixedSize(200, 55)
        icon = QIcon("img/registro-alumno.png")
        registrar_button.setIcon(icon)
        registrar_button.setIconSize(QSize(25,25))
        registrar_button.clicked.connect(self.record)
        
        update_button = QPushButton(" ACTUALIZAR", self)
        update_button.setStyleSheet(style.estilo)
        update_button.setCursor(Qt.CursorShape.PointingHandCursor)
        update_button.setFixedSize(200, 55)
        icon2 = QIcon("img/actualizar-alumno.png")
        update_button.setIcon(icon2)
        update_button.setIconSize(QSize(25,25))
        update_button.clicked.connect(self.update)
        
        delete_button = QPushButton(" ELIMINAR", self)
        delete_button.setCursor(Qt.CursorShape.PointingHandCursor)
        delete_button.setStyleSheet(style.estilo)
        delete_button.setFixedSize(200, 55)
        icon3 = QIcon("img/eliminar-alumno.png")
        delete_button.setIcon(icon3)
        delete_button.setIconSize(QSize(25,25))
        delete_button.clicked.connect(self.deleteRecord)
        
        actividad_button = QPushButton(" DISCIPLINA", self)
        actividad_button.setCursor(Qt.CursorShape.PointingHandCursor)
        actividad_button.setStyleSheet(style.estilo)
        actividad_button.setFixedSize(200, 55)
        icon4 = QIcon("img/disciplina.png")
        actividad_button.setIcon(icon4)
        actividad_button.setIconSize(QSize(25,25))
        actividad_button.clicked.connect(self.activity)
        
        pagos_button = QPushButton(" PAGOS", self)
        pagos_button.setCursor(Qt.CursorShape.PointingHandCursor)
        pagos_button.setStyleSheet(style.estilo)
        pagos_button.setFixedSize(200, 55)
        iconcobrar = QIcon("img/cobrar.png")
        pagos_button.setIcon(iconcobrar)
        pagos_button.setIconSize(QSize(40,40))
        pagos_button.clicked.connect(self.pagos)
        
        empleados_button = QPushButton(" EMPLEADOS", self)
        empleados_button.setCursor(Qt.CursorShape.PointingHandCursor)
        empleados_button.setStyleSheet(style.estilo)
        empleados_button.setFixedSize(200, 55)
        icon7 = QIcon("img/empleado.png")
        empleados_button.setIcon(icon7)
        empleados_button.setIconSize(QSize(25,25))
        empleados_button.clicked.connect(self.empleados)
        
        self.asistencia_button = QPushButton(" ASISTENCIA", self)
        self.asistencia_button.setCursor(Qt.CursorShape.PointingHandCursor)
        self.asistencia_button.setStyleSheet(style.estilo)
        self.asistencia_button.setFixedSize(200, 55)
        icon6 = QIcon("img/asistencia-alumno.png")
        self.asistencia_button.setIcon(icon6)
        self.asistencia_button.setIconSize(QSize(25,25))
        self.asistencia_button.clicked.connect(self.assistance)
        
        balances_button = QPushButton(" BALANCES", self)
        balances_button.setCursor(Qt.CursorShape.PointingHandCursor)
        balances_button.setStyleSheet(style.estilo)
        icon5 = QIcon("img/balance.png")
        balances_button.setIcon(icon5)
        balances_button.setIconSize(QSize(30,30))
        balances_button.setFixedSize(200, 55)
        balances_button.clicked.connect(self.balances)
        
        gastos_button = QPushButton(" CONTABILIDAD", self)
        gastos_button.setCursor(Qt.CursorShape.PointingHandCursor)
        gastos_button.setStyleSheet(style.estilo)
        gastos_button.setFixedSize(200, 55)
        icon8 = QIcon("img/contabilidad.png")
        gastos_button.setIcon(icon8)
        gastos_button.setIconSize(QSize(30,30))
        gastos_button.clicked.connect(self.registro_de_ingYegreso)
        
        # Crear un frame para contener los botones y la imagen
        frame = QFrame()
        frame.setStyleSheet("background-color: black;")
        frame_layout = QVBoxLayout(frame)
        
        # Agregar los botones al frame
        frame_layout.addWidget(logo_label)
        frame_layout.setAlignment(logo_label, Qt.AlignmentFlag.AlignHCenter)
        frame_layout.addWidget(registrar_button)
        frame_layout.setSpacing(20)
        frame_layout.addWidget(update_button)
        frame_layout.setSpacing(20)
        frame_layout.addWidget(delete_button)
        frame_layout.setSpacing(20)
        frame_layout.addWidget(actividad_button)
        frame_layout.setSpacing(20)
        frame_layout.addWidget(pagos_button)
        frame_layout.setSpacing(20)
        frame_layout.addWidget(empleados_button)
        frame_layout.setSpacing(20)
        frame_layout.addWidget(balances_button)
        frame_layout.setSpacing(20)
        frame_layout.addWidget(gastos_button)
        frame_layout.setSpacing(20)
        frame_layout.addWidget(self.asistencia_button)
        
        # Agregar el frame al QVBoxLayout
        layout_vertical1.addWidget(frame)
        
        # Crea un TAB(Pestaña)
        self.tab = QTabWidget()
        self.tab.setCursor(Qt.CursorShape.PointingHandCursor)
        self.tab.setStyleSheet(style.estilo_tab)
        self.tab.setUsesScrollButtons(True)
               
        # Crea las pestañas 
        pestania_record = QWidget()
        pestania_record.setStyleSheet("background-color: #f48c06;")
        pestania_updateRecord = QWidget()
        pestania_updateRecord.setStyleSheet("background-color: #f48c06;")
        pestania_deleteRecord = QWidget()
        pestania_deleteRecord.setStyleSheet("background-color: #f48c06;")
        pestania_actividad = QWidget()
        pestania_actividad.setStyleSheet("background-color: #f48c06;")
        pestania_pagos = QWidget()
        pestania_pagos.setStyleSheet("background-color: #f48c06;")
        pestania_view = QWidget()
        pestania_view.setStyleSheet("background-color: #f48c06;")
        pestania_empleados = QWidget()
        pestania_empleados.setStyleSheet("background-color: #f48c06;")
        pestania_resumen = QWidget()
        pestania_resumen.setStyleSheet("background-color: #f48c06;")
        
        # Agrega pestañas a cada Widget
        self.tab.addTab(pestania_record, 'REGISTRAR')
        self.tab.addTab(pestania_updateRecord, 'ACTUALIZAR')
        self.tab.addTab(pestania_deleteRecord, 'ELIMINAR')
        self.tab.addTab(pestania_actividad, 'DISCIPLINA')
        self.tab.addTab(pestania_pagos, 'PAGOS')
        self.tab.addTab(pestania_empleados, 'EMPLEADOS')
        self.tab.addTab(pestania_view, 'BALANCE')
        self.tab.addTab(pestania_resumen, 'CONTABILIDAD')
        
        self.tab.currentChanged.connect(self.onChangeTab)
        
        # ----------------------------------------------------
        # PESTAÑA REGISTRAR
        # SE CREA ComboBox 'RECORD'
        customer_details = QGroupBox("DETALLE DEL ALUMNO", pestania_record)
        customer_details.setStyleSheet(style.estilo_grupo)
        
        #Colocar el ComboBox a la grilla
        grid1 = QGridLayout(customer_details)

        # CONTENEDOR DE LOS LAYOUT HORIZONTALES
        layout_V1 = QVBoxLayout()
 
        # Crear layout horizontales para elementos '''LABEL y QLineEdit''' al diseño del QGroupBoxel QGroupBox
        layout_H1 = QHBoxLayout()
        layout_H1.setAlignment(Qt.AlignmentFlag.AlignLeft)
        layout_H2 = QHBoxLayout()
        layout_H2.setAlignment(Qt.AlignmentFlag.AlignLeft)
        layout_H = QHBoxLayout()
        layout_H.setAlignment(Qt.AlignmentFlag.AlignLeft)
            
        nombre1 = QLabel('Nombre:',customer_details)
        nombre1.setStyleSheet(style.label)
        nombre1.setFixedWidth(80)
        self.input_nombre1 = QLineEdit(customer_details)
        self.input_nombre1.setFixedWidth(200)
        self.input_nombre1.setPlaceholderText("Ingrese nombre...")
        self.input_nombre1.setStyleSheet(style.estilo_lineedit)
        self.input_nombre1.setFocus()
        layout_H1.addWidget(nombre1)        
        layout_H1.addWidget(self.input_nombre1)
        
        # Conexión a la base de datos MySQL
        conn = conectar_base_de_datos()
        cursor = conn.cursor()

        # Consulta para obtener los datos de una columna específica
        cursor.execute("SELECT nombre FROM usuario")
        datos = cursor.fetchall()
        sugerencia = [str(item[0]) for item in datos]

        lista_nombre = QCompleter(sugerencia)
        lista_nombre.setCaseSensitivity(Qt.CaseSensitivity.CaseInsensitive)
        lista_nombre.setCompletionMode(QCompleter.CompletionMode.PopupCompletion)
        lista_nombre.popup().setStyleSheet(style.completer)
        self.input_nombre1.setCompleter(lista_nombre)
        
        cursor.close()
        conn.close()
        
        apellido1 = QLabel('Apellido:',customer_details)
        apellido1.setStyleSheet(style.label)
        apellido1.setFixedWidth(80)
        self.input_apellido1 = QLineEdit(customer_details)
        self.input_apellido1.setStyleSheet(style.estilo_lineedit)
        self.input_apellido1.setPlaceholderText("Ingrese apellido...")
        self.input_apellido1.setFixedWidth(200)
        layout_H1.addWidget(apellido1)       
        layout_H1.addWidget(self.input_apellido1)
        
        dni = QLabel('DNI:',customer_details)
        dni.setStyleSheet(style.label)
        dni.setFixedWidth(100)
        self.input_dni = QLineEdit(customer_details)
        self.input_dni.setStyleSheet(style.estilo_lineedit)
        self.input_dni.setFixedWidth(200)
        self.input_dni.setPlaceholderText("Sin puntos")
        self.input_dni.setMaxLength(8)
        layout_H1.addWidget(dni)         
        layout_H1.addWidget(self.input_dni)

        sex = QLabel('Sexo:',customer_details)
        sex.setStyleSheet(style.label)
        sex.setFixedWidth(80)
        self.input_sex = QComboBox(customer_details)
        self.input_sex.setFixedWidth(200)
        self.input_sex.addItem("Seleccione...", "")  # Primer ítem vacío
        self.input_sex.addItem("Hombre", "Hombre")
        self.input_sex.addItem("Mujer", "Mujer")
        self.input_sex.setCurrentIndex(0)
        self.input_sex.setStyleSheet(style.estilo_combo)
        layout_H2.addWidget(sex)         
        layout_H2.addWidget(self.input_sex)
                
        age = QLabel('Edad:',customer_details)
        age.setStyleSheet(style.label)
        age.setFixedWidth(80)
        self.input_age = QLineEdit(customer_details)
        self.input_age.setStyleSheet(style.estilo_lineedit)
        self.input_age.setFixedWidth(200)
        self.input_age.setPlaceholderText("23")
        self.input_age.setMaxLength(2)
        layout_H2.addWidget(age)         
        layout_H2.addWidget(self.input_age)
               
        celular = QLabel('N° celular:', customer_details)
        celular.setStyleSheet(style.label)
        celular.setFixedWidth(100)
        self.input_celular = QLineEdit(customer_details)
        self.input_celular.setStyleSheet(style.estilo_lineedit)
        self.input_celular.setFixedWidth(200)
        self.input_celular.setPlaceholderText("Ej: 3424789123")
        layout_H2.addWidget(celular)
        layout_H2.addWidget(self.input_celular)

        date = QLabel('Fecha:',customer_details)
        date.setStyleSheet(style.label)
        date.setFixedWidth(80)
        self.input_date = QDateEdit(customer_details)
        self.input_date.setStyleSheet(style.fecha)
        self.input_date.setFixedWidth(200)
        self.input_date.setLocale(QLocale("es-AR"))
        self.input_date.setCursor(Qt.CursorShape.PointingHandCursor)
        self.input_date.setCalendarPopup(True)
        self.input_date.setDate(QDate.currentDate())
        self.input_date.setDisplayFormat("dd/MM/yyyy")
        layout_H.addWidget(date)      
        layout_H.addWidget(self.input_date)
        
        # Crear un layout horizontal para los botones y Se coloca el icono de la 'manito' al cursor
        botonera_registro = QHBoxLayout()
        botonera_registro.setAlignment(Qt.AlignmentFlag.AlignRight)
        button_Guardar = QPushButton('GUARDAR',customer_details)
        button_Guardar.setFixedWidth(250)
        button_Guardar.setCursor(Qt.CursorShape.PointingHandCursor)
        button_Guardar.setStyleSheet(style.estilo_boton)
        button_search = QPushButton('BUSCAR',customer_details)
        button_search.setFixedWidth(250)
        button_search.setCursor(Qt.CursorShape.PointingHandCursor)
        button_search.setStyleSheet(style.estilo_boton)
        
        # Crear un layout horizontal para los botones
        botonera_registro2 = QHBoxLayout()
        botonera_registro2.setAlignment(Qt.AlignmentFlag.AlignRight)
        button_MostrarTabla = QPushButton('MOSTRAR TABLA',customer_details)
        button_MostrarTabla.setFixedWidth(250)
        button_MostrarTabla.setCursor(Qt.CursorShape.PointingHandCursor)
        button_MostrarTabla.setStyleSheet(style.estilo_boton)
        button_limparTabla = QPushButton('LIMPIAR TABLA',customer_details)
        button_limparTabla.setFixedWidth(250)
        button_limparTabla.setCursor(Qt.CursorShape.PointingHandCursor)
        button_limparTabla.setStyleSheet(style.estilo_boton)
        
        botonera_registro3 = QHBoxLayout()
        botonera_registro3.setAlignment(Qt.AlignmentFlag.AlignRight)
        excel_registro = QPushButton('DESCARGAR PLANILLA', customer_details)
        excel_registro.setFixedWidth(250)
        excel_registro.setCursor(Qt.CursorShape.PointingHandCursor)
        excel_registro.setStyleSheet(style.boton_excel)
        
        # Establece los botones a los layout horizontales
        botonera_registro.addWidget(button_Guardar)
        botonera_registro.addWidget(button_search)
        botonera_registro2.addWidget(button_MostrarTabla)
        botonera_registro2.addWidget(button_limparTabla)
        botonera_registro3.addWidget(excel_registro)
        
        # Agregar los layouts horizontales al layout vertical
        layout_cont1 = QHBoxLayout()
        layout_cont2 = QHBoxLayout()
        layout_cont3 = QHBoxLayout()
        
        layout_cont1.addLayout(layout_H1)
        layout_cont1.addLayout(botonera_registro)
        layout_cont2.addLayout(layout_H2)
        layout_cont2.addLayout(botonera_registro2)
        layout_cont3.addLayout(layout_H)
        layout_cont3.addLayout(botonera_registro3)
        
        layout_V1.addLayout(layout_cont1)
        layout_V1.addLayout(layout_cont2)
        layout_V1.addLayout(layout_cont3)
        
        # Crear un QGridLayout para organizar los elementos en la QGroupBox
        grid1.addLayout(layout_V1,0,0,1,1)
        
        # Crear una QTableWidget
        self.tablaRecord = QTableWidget()
        self.tablaRecord.setCursor(Qt.CursorShape.PointingHandCursor)
        self.tablaRecord.setStyleSheet(style.esttabla)
        grid1.addWidget(self.tablaRecord,1,0,1,1)
        
        # conecta las señales a las funciones
        button_Guardar.clicked.connect(self.guardar)
        button_search.clicked.connect(self.search)
        button_MostrarTabla.clicked.connect(self.mostrarTabla)
        button_limparTabla.clicked.connect(self.claer_tabla)
        excel_registro.clicked.connect(self.tabla_registro)
        
        # Establecer el diseño del QGroupBox
        customer_details.setLayout(grid1)
        
        # Agregar el QGroupBox a la primera pestaña (tab1)
        tab1_layout = QVBoxLayout()
        tab1_layout.addWidget(customer_details)
        pestania_record.setLayout(tab1_layout)
        
        #-----------------------------------------------------------------
        # CREAR LA PESTAÑA 'ACTUALIZAR REGISTRO'
        # CREA UN GROUPBOX 
        update_customer_details = QGroupBox("DETALLE DEL ALUMNO", pestania_updateRecord)
        update_customer_details.setStyleSheet(style.estilo_grupo)
        
        # CONTENEDOR DE LOS LAYOUT HORIZONTALES
        layout_V2 = QVBoxLayout()
 
        # Crear un diseño para elementos '''LABEL y QLineEdit''' al diseño del QGroupBoxel QGroupBox
        layout_ele1 = QHBoxLayout()
        layout_ele1.setAlignment(Qt.AlignmentFlag.AlignLeft)
        layout_ele2 = QHBoxLayout()
        layout_ele2.setAlignment(Qt.AlignmentFlag.AlignLeft)
        layout_ele3 = QHBoxLayout()
        layout_ele3.setAlignment(Qt.AlignmentFlag.AlignLeft)
                
        nombre2 = QLabel('Nombre:',update_customer_details)
        nombre2.setStyleSheet(style.label)
        nombre2.setFixedWidth(80)
        self.input_nombre2 = QLineEdit(update_customer_details)
        self.input_nombre2.setStyleSheet(style.estilo_lineedit)
        self.input_nombre2.setPlaceholderText("Ingrese nombre...")
        # self.input_nombre2.setValidator(validator)
        self.input_nombre2.setFixedWidth(200)
        self.input_nombre2.setFocus()
        layout_ele1.addWidget(nombre2)       
        layout_ele1.addWidget(self.input_nombre2)
        
        # Conexión a la base de datos MySQL
        conn = conectar_base_de_datos()
        cursor = conn.cursor()

        # Consulta para obtener los datos de una columna específica
        cursor.execute("SELECT nombre FROM usuario")
        datos = cursor.fetchall()
        suger = [str(item[0]) for item in datos]

        lista_nombres = QCompleter(suger)
        lista_nombres.setCaseSensitivity(Qt.CaseSensitivity.CaseInsensitive)
        lista_nombres.setCompletionMode(QCompleter.CompletionMode.PopupCompletion)
        lista_nombres.popup().setStyleSheet(style.completer)
        self.input_nombre2.setCompleter(lista_nombres)
        
        cursor.close()
        conn.close()
        
        apellido2 = QLabel('Apellido:',update_customer_details)
        apellido2.setStyleSheet(style.label)
        apellido2.setFixedWidth(80)
        self.input_apellido2 = QLineEdit(update_customer_details)
        self.input_apellido2.setStyleSheet(style.estilo_lineedit)
        self.input_apellido2.setPlaceholderText("Ingrese apellido...")
        self.input_apellido2.setEnabled(False)
        self.input_apellido2.setFixedWidth(200)
        layout_ele1.addWidget(apellido2)          
        layout_ele1.addWidget(self.input_apellido2)
        
        dni2 = QLabel('DNI:',update_customer_details)
        dni2.setStyleSheet(style.label)
        dni2.setFixedWidth(100)
        self.input_dni2 = QLineEdit(update_customer_details)
        self.input_dni2.setStyleSheet(style.estilo_lineedit)
        self.input_dni2.setEnabled(False)
        self.input_dni2.setFixedWidth(200)
        self.input_dni2.setPlaceholderText("Sin puntos")
        self.input_dni2.setMaxLength(8)
        layout_ele1.addWidget(dni2)    
        layout_ele1.addWidget(self.input_dni2)

        sex2 = QLabel('Sexo:',update_customer_details)
        sex2.setStyleSheet(style.label)
        sex2.setFixedWidth(80)
        self.input_sex2 = QComboBox(update_customer_details)
        self.input_sex2.setStyleSheet(style.estilo_combo)
        self.input_sex2.setFixedWidth(200)
        self.input_sex2.setEnabled(False)
        self.input_sex2.addItem("Seleccione...", "")  # Primer ítem vacío
        self.input_sex2.addItem("Hombre", "Hombre")
        self.input_sex2.addItem("Mujer", "Mujer")
        self.input_sex2.setCurrentIndex(0)
        layout_ele2.addWidget(sex2)       
        layout_ele2.addWidget(self.input_sex2)
        
        age2 = QLabel('Edad:',update_customer_details)
        age2.setStyleSheet(style.label)
        age2.setFixedWidth(80)
        self.input_age2 = QLineEdit(update_customer_details)
        self.input_age2.setStyleSheet(style.estilo_lineedit)
        self.input_age2.setEnabled(False)
        self.input_age2.setPlaceholderText("23")
        self.input_age2.setFixedWidth(200)
        self.input_age2.setMaxLength(2)
        layout_ele2.addWidget(age2)      
        layout_ele2.addWidget(self.input_age2)
               
        celular2 = QLabel('N° celular:', update_customer_details)
        celular2.setStyleSheet(style.label)
        celular2.setFixedWidth(100)
        self.input_celular2 = QLineEdit(update_customer_details)
        self.input_celular2.setStyleSheet(style.estilo_lineedit)
        self.input_celular2.setPlaceholderText("Ej: 3424789123")
        self.input_celular2.setEnabled(False)
        self.input_celular2.setFixedWidth(200)
        layout_ele2.addWidget(celular2)       
        layout_ele2.addWidget(self.input_celular2)
        
        date2 = QLabel('Fecha:', update_customer_details)
        date2.setStyleSheet(style.label)
        date2.setFixedWidth(80)
        self.input_date2 = QDateEdit(update_customer_details)
        self.input_date2.setLocale(QLocale("es-AR"))
        self.input_date2.setStyleSheet(style.fecha)
        self.input_date2.setCursor(Qt.CursorShape.PointingHandCursor)
        self.input_date2.setEnabled(False)
        self.input_date2.setFixedWidth(200)
        self.input_date2.setCalendarPopup(True)
        self.input_date2.setDate(QDate.currentDate())
        self.input_date2.setDisplayFormat("dd/MM/yyyy")
        layout_ele3.addWidget(date2)     
        layout_ele3.addWidget(self.input_date2)
            
        # Crear un layout horizontal para los botones y Se coloca el icono de la 'manito' al cursor
        button_layout2 = QHBoxLayout()
        button_layout2.setAlignment(Qt.AlignmentFlag.AlignRight)
        button_Buscar1 = QPushButton('BUSCAR', update_customer_details)
        button_Buscar1.setFixedWidth(250)
        button_Buscar1.setCursor(Qt.CursorShape.PointingHandCursor)
        button_Buscar1.setStyleSheet(style.estilo_boton)
        button_Mostrar_tabla = QPushButton('MOSTRAR TABLA',update_customer_details)        
        button_Mostrar_tabla.setFixedWidth(250)
        button_Mostrar_tabla.setCursor(Qt.CursorShape.PointingHandCursor)
        button_Mostrar_tabla.setStyleSheet(style.estilo_boton)
        
        # AGREGAR AL "LAYOUT"
        button_layout2.addWidget(button_Buscar1)
        button_layout2.addWidget(button_Mostrar_tabla)
        layout_ele1.addLayout(button_layout2)
        
        button_layout3 = QHBoxLayout()
        button_layout3.setAlignment(Qt.AlignmentFlag.AlignRight)
        button_Actualizar = QPushButton('ACTUALIZAR',update_customer_details)
        button_Actualizar.setFixedWidth(250)
        button_Actualizar.setCursor(Qt.CursorShape.PointingHandCursor)
        button_Actualizar.setStyleSheet(style.estilo_boton)
        button_Limpiar = QPushButton('LIMPIAR',update_customer_details)
        button_Limpiar.setFixedWidth(250)
        button_Limpiar.setCursor(Qt.CursorShape.PointingHandCursor)
        button_Limpiar.setStyleSheet(style.estilo_boton)
        
        # AGREGAR AL "LAYOUT"
        button_layout3.addWidget(button_Actualizar)
        button_layout3.addWidget(button_Limpiar)
        layout_ele2.addLayout(button_layout3)
        
        # CONECCIONES A FUNCIONES
        button_Mostrar_tabla.clicked.connect(self.ver)
        button_Actualizar.clicked.connect(self.actualizar)
        button_Buscar1.clicked.connect(self.buscar)
        button_Limpiar.clicked.connect(self.limpiar)
        
        # Agregar los layouts horizontales al layout vertical
        layout_V2.addLayout(layout_ele1)
        layout_V2.addLayout(layout_ele2)
        layout_V2.addLayout(layout_ele3)
        
        # Crear un QGridLayout para organizar los elementos en la QGroupBox
        grid2 = QGridLayout(update_customer_details)
        grid2.addLayout(layout_V2,0,0,1,4)
        
        # Crear una QTableWidget
        self.tablaUpdateRecord = QTableWidget(update_customer_details)
        self.tablaUpdateRecord.setCursor(Qt.CursorShape.PointingHandCursor)
        self.tablaUpdateRecord.setStyleSheet(style.esttabla)
        
        # PARA TABLA DE PESTAÑA ACTUALIZAR
        self.tablaUpdateRecord.clicked.connect(self.seleccionYautoCompletado)

        grid2.addWidget(self.tablaUpdateRecord,1,0,1,4)
        
        # Establecer el diseño del QGroupBox
        update_customer_details.setLayout(grid2)
        
        # Agregar el QGroupBox a la primera pestaña (tab1)
        tab2_layout = QVBoxLayout()
        tab2_layout.addWidget(update_customer_details)
        pestania_updateRecord.setLayout(tab2_layout)
        
        #-----------------------------------------------------------------
        # PESTAÑA ELIMINAR
        # CREA EL GRUPOBOX
        delete_Record = QGroupBox("ELIMINACION DE ALUMNO", pestania_deleteRecord)
        delete_Record.setStyleSheet(style.estilo_grupo)
        
        # ESTABLECE EL COMBOBOX A LA GRILLA
        grid3 = QGridLayout(delete_Record)

        vert = QHBoxLayout()
        
        # SE CREA UN Diseño de formulario PARA LOS ELEMENTOS
        layout_H7 = QHBoxLayout()
        nombre_buscar3 = QLabel('Nombre:',delete_Record)
        nombre_buscar3.setStyleSheet(style.label)
        nombre_buscar3.setFixedWidth(100)
        self.nombre_buscar3 = QLineEdit(delete_Record)
        self.nombre_buscar3.setStyleSheet(style.estilo_lineedit)
        self.nombre_buscar3.setFixedWidth(300)
        self.nombre_buscar3.setPlaceholderText("Ingrese el nombre o su inicial")
        self.nombre_buscar3.setFocus()
        layout_H7.addWidget(nombre_buscar3)
        layout_H7.addWidget(self.nombre_buscar3)
        
        # Conexión a la base de datos MySQL
        conn = conectar_base_de_datos()
        cursor = conn.cursor()

        # Consulta para obtener los datos de una columna específica
        cursor.execute("SELECT nombre FROM usuario WHERE habilitado = 1")
        datos = cursor.fetchall()
        suger2 = [str(item[0]) for item in datos]

        lista_nombre = QCompleter(suger2)
        lista_nombre.setCaseSensitivity(Qt.CaseSensitivity.CaseInsensitive)
        lista_nombre.setCompletionMode(QCompleter.CompletionMode.PopupCompletion)
        lista_nombre.popup().setStyleSheet(style.completer)
        self.nombre_buscar3.setCompleter(lista_nombre)
        
        cursor.close()
        conn.close()
            
        # Agregar al "layout" botones
        layout_horiz2 = QHBoxLayout()
        layout_horiz2.setAlignment(Qt.AlignmentFlag.AlignRight)
        button_Buscar = QPushButton('BUSCAR',delete_Record)
        button_Buscar.setFixedWidth(250)
        button_Buscar.setCursor(Qt.CursorShape.PointingHandCursor)
        button_Buscar.setStyleSheet(style.estilo_boton)
        button_Eliminar = QPushButton('ELIMINAR',delete_Record)
        button_Eliminar.setFixedWidth(250)
        button_Eliminar.setCursor(Qt.CursorShape.PointingHandCursor)
        button_Eliminar.setStyleSheet(style.estilo_boton)
        button_LimpiarTabla = QPushButton('LIMPIAR TABLA',delete_Record)
        button_LimpiarTabla.setFixedWidth(250)
        button_LimpiarTabla.setCursor(Qt.CursorShape.PointingHandCursor)
        button_LimpiarTabla.setStyleSheet(style.estilo_boton)
        
        # Agregar al "layout" botones
        layout_horiz2.addWidget(button_Buscar)
        layout_horiz2.addWidget(button_Eliminar)
        layout_horiz2.addWidget(button_LimpiarTabla)
        layout_horiz2.addSpacing(15)
        
        vert.addLayout(layout_H7)
        vert.addLayout(layout_horiz2)
        
        # Conexiones de selañes a las funciones
        button_Buscar.clicked.connect(self.buscar_para_eliminar)
        button_Eliminar.clicked.connect(self.delete)
        button_LimpiarTabla.clicked.connect(self.limpiar_tabla)
        
        # Agrega el layout en la grid
        grid3.addLayout(vert,0,0,1,1)
        
        self.tablaDeleteRecord = QTableWidget(delete_Record)
        self.tablaDeleteRecord.setCursor(Qt.CursorShape.PointingHandCursor)
        self.tablaDeleteRecord.setStyleSheet(style.esttabla)
        self.tablaDeleteRecord.clicked.connect(self.obtenerDato)
        
        # Agrega la tabla al grid
        grid3.addWidget(self.tablaDeleteRecord,1,0,1,1)
        
        # Establecer el diseño del QGroupBox
        delete_Record.setLayout(grid3)
        
        # Agregar el QGroupBox a la primera pestaña (tab1)
        tab3_layout = QVBoxLayout()
        tab3_layout.addWidget(delete_Record)
        pestania_deleteRecord.setLayout(tab3_layout)
               
        #-----------------------------------------------------------------
        # PESTAÑA DE DISCIPLINA
        # CREA EL COMBOBOX
        comboActiv = QGroupBox("DETALLE DEL DISCIPLINA", pestania_actividad)
        comboActiv.setStyleSheet(style.estilo_grupo)
        
        # ESTABLECE EL COMBOBOX A LA GRILLA
        grid4 = QGridLayout(comboActiv)

        # CREA LAYOUT HORIZONTAL PARA LOS ELEMENTOS 
        layout_H8 = QHBoxLayout()
        layout_H8.setAlignment(Qt.AlignmentFlag.AlignLeft)
        layout_H9 = QHBoxLayout()
        layout_H9.setAlignment(Qt.AlignmentFlag.AlignLeft)
          
        disciplina4 = QLabel('Disciplina:',comboActiv)
        disciplina4.setStyleSheet(style.label)
        disciplina4.setFixedWidth(120)
        self.input_disciplina4 = QLineEdit(comboActiv)
        self.input_disciplina4.setStyleSheet(style.estilo_lineedit)
        self.input_disciplina4.setFixedWidth(200)
        self.input_disciplina4.setPlaceholderText("Ingrese una disciplina")
        self.input_disciplina4.setFocus()
        layout_H8.addWidget(disciplina4)        
        layout_H8.addWidget(self.input_disciplina4)
        
        precio = QLabel('Precio($):',comboActiv)
        precio.setStyleSheet(style.label)
        precio.setFixedWidth(120)
        self.input_precio = QLineEdit(comboActiv)
        self.input_precio.setStyleSheet(style.estilo_lineedit)
        self.input_precio.setFixedWidth(200)
        self.input_precio.setPlaceholderText("Ej: 5000")
        layout_H9.addWidget(precio)     
        layout_H9.addWidget(self.input_precio)  
                
        # layout horizontal para botones
        layout_botones9_10 = QHBoxLayout()
        layout_botones9_10.setAlignment(Qt.AlignmentFlag.AlignRight)
        button_Gurd = QPushButton('GUARDAR',comboActiv)
        button_Gurd.setFixedWidth(200)
        button_Gurd.setCursor(Qt.CursorShape.PointingHandCursor)
        button_Gurd.setStyleSheet(style.estilo_boton)
        button_Tabla = QPushButton('MOSTRAR TABLA',comboActiv)
        button_Tabla.setFixedWidth(200)
        button_Tabla.setCursor(Qt.CursorShape.PointingHandCursor)
        button_Tabla.setStyleSheet(style.estilo_boton)
        button_Actaul = QPushButton('ACTUALIZAR',comboActiv)
        button_Actaul.setFixedWidth(200)
        button_Actaul.setCursor(Qt.CursorShape.PointingHandCursor)
        button_Actaul.setStyleSheet(style.estilo_boton)
        
        # layout horizontal para botones
        layout_botones11_12 = QHBoxLayout()
        layout_botones11_12.setAlignment(Qt.AlignmentFlag.AlignRight)
        button_Elim = QPushButton('ELIMINAR',comboActiv)
        button_Elim.setFixedWidth(200)
        button_Elim.setCursor(Qt.CursorShape.PointingHandCursor)
        button_Elim.setStyleSheet(style.estilo_boton)
        limpiar_tabla = QPushButton('LIMPIAR TABLA',comboActiv)
        limpiar_tabla.setFixedWidth(200)
        limpiar_tabla.setCursor(Qt.CursorShape.PointingHandCursor)
        limpiar_tabla.setStyleSheet(style.estilo_boton)
        button_limp = QPushButton('LIMPIAR CAMPOS',comboActiv)
        button_limp.setFixedWidth(200)
        button_limp.setCursor(Qt.CursorShape.PointingHandCursor)
        button_limp.setStyleSheet(style.estilo_boton)
        
        lay_excel = QHBoxLayout()
        lay_excel.setAlignment(Qt.AlignmentFlag.AlignRight)
        excel = QPushButton('DESCARGAR PLANILLA', comboActiv)
        excel.setFixedWidth(200)
        excel.setCursor(Qt.CursorShape.PointingHandCursor)
        excel.setStyleSheet(style.boton_excel)
        
        # AGREGA A LOS "LAYOUT"
        layout_botones9_10.addWidget(button_Gurd)
        layout_botones9_10.addWidget(button_Tabla)
        layout_botones9_10.addWidget(button_Actaul)
        layout_botones11_12.addWidget(limpiar_tabla)
        layout_botones11_12.addWidget(button_Elim)
        layout_botones11_12.addWidget(button_limp)
        lay_excel.addWidget(excel)
        
        layout_grup1 = QHBoxLayout()
        layout_grup1.addLayout(layout_H8)
        layout_grup1.addLayout(layout_botones9_10)
        layout_grup2 = QHBoxLayout()
        layout_grup2.addLayout(layout_H9)
        layout_grup2.addLayout(layout_botones11_12)
        layout_grup3 = QHBoxLayout()
        layout_grup3.addLayout(lay_excel)
        
        
        # LAYOUT VENTICAL Y AGREGA LOS LAYOUT HORIZONTALES
        layoutV = QVBoxLayout()
        layoutV.addLayout(layout_grup1)
        layoutV.addLayout(layout_grup2)
        layoutV.addLayout(layout_grup3)        
       
        # ESTABLECE EL LAYOUT VENTICAL A LA GRILLA
        grid4.addLayout(layoutV,0,0,1,1)
            
        # CONECCION LAS SEÑALES A LAS FUNCIONES
        button_Gurd.clicked.connect(self.guardarACTIV)
        button_Tabla.clicked.connect(self.mostrarACTIC)
        button_Actaul.clicked.connect(self.actualizarACTIV)
        button_Elim.clicked.connect(self.eliminarACTIV)
        button_limp.clicked.connect(self.limp)
        limpiar_tabla.clicked.connect(self.limpiar_tabla_disciplina)
        excel.clicked.connect(self.planilla_disciplina)

        # CREA LA TABLA
        self.tableActivi = QTableWidget(comboActiv)
        self.tableActivi.setCursor(Qt.CursorShape.PointingHandCursor)
        self.tableActivi.setStyleSheet(style.esttabla)
        
        # PARA TABLA DE PERSTAÑA DISCIPLINA 'ACTUALIZAR'
        self.tableActivi.clicked.connect(self.seleccionDeDatos)
        self.tableActivi.itemSelectionChanged.connect(self.seleccionar_fila)
        
        # ESTABLECE LA TABLA A LA GRILLA
        grid4.addWidget(self.tableActivi,1,0,1,1)
                
        # Establecer el diseño del QGroupBox
        comboActiv.setLayout(grid4)
        
        # Agregar el QGroupBox a la primera pestaña (tab1)
        tab4_layout = QVBoxLayout()
        tab4_layout.addWidget(comboActiv)
        pestania_actividad.setLayout(tab4_layout)
        
        #-----------------------------------------------------------------
        # PESTAÑA DE PAGOS
        # CREA EL GrupoBOX
        grupo_pagos = QGroupBox("DETALLE DEL PAGO", pestania_pagos)
        grupo_pagos.setStyleSheet(style.estilo_grupo)
        
        # ESTABLECE EL COMBOBOX A LA GRILLA
        gr = QGridLayout(grupo_pagos)
        
        v = QVBoxLayout()
        
        # CREA LAYOUT HORIZONTAL PARA LOS ELEMENTOS 
        layout_elementos_pagos = QHBoxLayout()
        layout_elementos_pagos.setAlignment(Qt.AlignmentFlag.AlignLeft)
        layout_elementos_pagos2 = QHBoxLayout()
        layout_elementos_pagos2.setAlignment(Qt.AlignmentFlag.AlignLeft)
        layout_elementos_pagos3 = QHBoxLayout()
        layout_elementos_pagos3.setAlignment(Qt.AlignmentFlag.AlignLeft)
        
        idUser = QLabel('DNI',grupo_pagos)
        idUser.setStyleSheet(style.label)
        idUser.setFixedWidth(140)
        self.idUser = QLineEdit(grupo_pagos)
        self.idUser.setStyleSheet(style.estilo_lineedit)
        self.idUser.setPlaceholderText("Ingrese un DNI")
        self.idUser.setFixedWidth(180)
        self.idUser.setFocus()
        layout_elementos_pagos.addWidget(idUser)     
        layout_elementos_pagos.addWidget(self.idUser)

        conn = conectar_base_de_datos()
        cursor = conn.cursor()

        # Consulta para obtener los datos de una columna específica
        cursor.execute("SELECT dni FROM usuario WHERE habilitado = 1 ORDER BY dni ASC")
        dato = cursor.fetchall()
        self.sugerencia = [str(item[0]) for item in dato]
        
        # Almacenar sugerencias en un conjunto para verificación rápida
        self.sugerencias_set = set(self.sugerencia)
        
        idDNI = QCompleter(self.sugerencia)
        idDNI.setCaseSensitivity(Qt.CaseSensitivity.CaseInsensitive)
        idDNI.setCompletionMode(QCompleter.CompletionMode.PopupCompletion)
        idDNI.popup().setStyleSheet(style.completer)
        self.idUser.setCompleter(idDNI)
        
        cursor.close()
        conn.close()
        
        # Inicializar current_id_disciplina con un valor por defecto, puede ser None o algún otro valor válido
        self.current_id_disciplina = None
        
        idDis = QLabel('Disciplina:',grupo_pagos)
        idDis.setStyleSheet(style.label)
        idDis.setFixedWidth(150)
        self.idDis = QLineEdit(grupo_pagos)
        self.idDis.setStyleSheet(style.estilo_lineedit)
        self.idDis.setFixedWidth(250)
        self.idDis.setPlaceholderText("Nombre de la disciplina")
        layout_elementos_pagos.addWidget(idDis)     
        layout_elementos_pagos.addWidget(self.idDis)
        
        self.label_monto = QLabel()
        self.label_monto.setStyleSheet(style.label)
        layout_elementos_pagos.addWidget(self.label_monto)
        
        #  Conexión a la base de datos MySQL
        conn = conectar_base_de_datos()
        cursor = conn.cursor()
        cursor.execute("SELECT id_disciplina, nombre, precio FROM disciplina WHERE disciplina.habilitado = 1 ORDER BY nombre ASC")
        dato = cursor.fetchall()
        self.disciplinas = {str(item[1]): (item[0], item[2]) for item in dato}  # Mapear nombre a (id_disciplina, precio)
        
        sugerencias = list(self.disciplinas.keys())
        
        # Almacenar sugerencias en un conjunto para verificación rápida
        self.sugeren_set = set(sugerencias)
        
        idDis = QCompleter(sugerencias)
        idDis.setCaseSensitivity(Qt.CaseSensitivity.CaseInsensitive)
        idDis.setCompletionMode(QCompleter.CompletionMode.PopupCompletion)
        idDis.popup().setStyleSheet(style.completer)
        self.idDis.setCompleter(idDis)
    
        cursor.close()
        conn.close()
        
        # Muestra el precio de cada disciplina al elegitr la disciplina
        idDis.activated.connect(self.actualizar_precio)
        
        fechaDePago = QLabel('Fecha de pago:',grupo_pagos)
        fechaDePago.setStyleSheet(style.label)
        fechaDePago.setFixedWidth(140)
        self.input_fechaDePago = QDateEdit(grupo_pagos)
        self.input_fechaDePago.setCursor(Qt.CursorShape.PointingHandCursor)
        self.input_fechaDePago.setLocale(QLocale("es-AR"))
        self.input_fechaDePago.setStyleSheet(style.fecha)
        self.input_fechaDePago.setFixedWidth(180)
        self.input_fechaDePago.setDate(QDate.currentDate()) 
        self.input_fechaDePago.setCalendarPopup(True)
        self.input_fechaDePago.setDisplayFormat("dd/MM/yyyy")
        layout_elementos_pagos2.addWidget(fechaDePago)        
        layout_elementos_pagos2.addWidget(self.input_fechaDePago)
        
        tipoDePago = QLabel('Medio de pago:',grupo_pagos)
        tipoDePago.setStyleSheet(style.label)
        tipoDePago.setFixedWidth(150)
        self.input_tipoDePago = QLineEdit(grupo_pagos)
        self.input_tipoDePago.setStyleSheet(style.estilo_lineedit)
        self.input_tipoDePago.setFixedWidth(300)
        self.input_tipoDePago.setPlaceholderText("'Efectivo' o 'Transferencia'")
        layout_elementos_pagos2.addWidget(tipoDePago)   
        layout_elementos_pagos2.addWidget(self.input_tipoDePago)
        
        medio_de_pago = ["Efectivo", "Transferencia"]
        completer2 = QCompleter(medio_de_pago)
        completer2.setCaseSensitivity(Qt.CaseSensitivity.CaseInsensitive)
        completer2.setCompletionMode(QCompleter.CompletionMode.PopupCompletion)
        completer2.popup().setStyleSheet(style.completer)
        self.input_tipoDePago.setCompleter(completer2)
        
        layoutBoton = QHBoxLayout() 
        layoutBoton.setAlignment(Qt.AlignmentFlag.AlignRight)
        guar_pagos = QPushButton('GUARDAR', grupo_pagos)
        guar_pagos.setCursor(Qt.CursorShape.PointingHandCursor)
        guar_pagos.setFixedWidth(250)
        guar_pagos.setStyleSheet(style.estilo_boton) 
        most_pagos = QPushButton('MOSTRAR', grupo_pagos)
        most_pagos.setFixedWidth(250)
        most_pagos.setCursor(Qt.CursorShape.PointingHandCursor)
        most_pagos.setStyleSheet(style.estilo_boton)
        layoutBoton.addWidget(guar_pagos)
        layoutBoton.addWidget(most_pagos)
        
        layoutBoton2 = QHBoxLayout() 
        layoutBoton2.setAlignment(Qt.AlignmentFlag.AlignRight)
        act_pagos = QPushButton('ACTUALIZAR', grupo_pagos)
        act_pagos.setCursor(Qt.CursorShape.PointingHandCursor)
        act_pagos.setFixedWidth(250)
        act_pagos.setStyleSheet(style.estilo_boton)
        eli_pagos = QPushButton('ELIMINAR', grupo_pagos)
        eli_pagos.setCursor(Qt.CursorShape.PointingHandCursor)
        eli_pagos.setFixedWidth(250)
        eli_pagos.setStyleSheet(style.estilo_boton)
        layoutBoton2.addWidget(act_pagos)
        layoutBoton2.addWidget(eli_pagos)
        
        layout_panilla = QHBoxLayout() 
        layout_panilla.setAlignment(Qt.AlignmentFlag.AlignRight)
        planilla = QPushButton('PLANILLA', grupo_pagos)
        planilla.setFixedWidth(250)
        planilla.setCursor(Qt.CursorShape.PointingHandCursor)
        planilla.setStyleSheet(style.boton_excel)
        layout_panilla.addWidget(planilla)
        
        l1 = QHBoxLayout()
        l1.addLayout(layout_elementos_pagos)
        l1.addLayout(layoutBoton)
        l2 = QHBoxLayout()
        l2.addLayout(layout_elementos_pagos2)
        l2.addLayout(layoutBoton2)
        l3 = QHBoxLayout()
        l3.addLayout(layout_elementos_pagos3)
        l3.addLayout(layout_panilla)
        
        # contenedor de los layout horizontales
        v.addLayout(l1)
        v.addLayout(l2)
        v.addLayout(l3)
        
        # Señales de botones
        guar_pagos.clicked.connect(self.guardarPagos)
        most_pagos.clicked.connect(self.mostrarPagos)
        act_pagos.clicked.connect(self.actualizarPagos)
        eli_pagos.clicked.connect(self.eliminarPagos)
        planilla.clicked.connect(self.planilla_pagos)
        
        # AGREGA AL "GRID"
        gr.addLayout(v,0,0,1,1)
        
        # CREA LA TABLA
        self.tablePagos = QTableWidget(grupo_pagos)
        self.tablePagos.setCursor(Qt.CursorShape.PointingHandCursor)
        self.tablePagos.setStyleSheet(style.esttabla)
        
        self.tablePagos.clicked.connect(self.establecer_datos)
        self.tablePagos.itemSelectionChanged.connect(self.seleccionar_fila)
        
        # ESTABLECE LA TABLA A LA GRILLA
        gr.addWidget(self.tablePagos,1,0,1,1)
        
        # Establecer el diseño del QGroupBox
        grupo_pagos.setLayout(gr)
        
        # Agregar el QGroupBox a la primera pestaña (tab1)
        tab_pagos_layout = QVBoxLayout()
        tab_pagos_layout.addWidget(grupo_pagos)
        pestania_pagos.setLayout(tab_pagos_layout)
        
        #-----------------------------------------------------------------
        # PESTAÑA DE BALANCES
        # CREA COMBOBOX 
        comboView = QGroupBox("DETALLE DEL BALANCES", pestania_view)
        comboView.setStyleSheet(style.estilo_grupo)
        
        # ESTABLECE EL COMBOBOX A LA GRILLA
        grid5 = QGridLayout(comboView)
        
        # LAYOUT VENTICAL
        layout_V4 = QVBoxLayout()
        
        # LAYOUT HORIZONTAL
        elementos = QHBoxLayout()
        elementos.setAlignment(Qt.AlignmentFlag.AlignLeft)
        elementos2 = QHBoxLayout()
        elementos2.setAlignment(Qt.AlignmentFlag.AlignLeft)
        elementos3 = QHBoxLayout()
        elementos3.setAlignment(Qt.AlignmentFlag.AlignLeft)
        
        view_nomb = QLabel("N° DNI:",comboView)
        view_nomb.setStyleSheet(style.label)
        view_nomb.setFixedWidth(140)
        self.view_nomb = QLineEdit(comboView)
        self.view_nomb.setPlaceholderText("Ingrese un DNI")
        self.view_nomb.setMaxLength(8)
        self.view_nomb.setFixedWidth(200)
        self.view_nomb.setFocus()
        self.view_nomb.setStyleSheet(style.estilo_lineedit)
        elementos.addWidget(view_nomb)     
        elementos.addWidget(self.view_nomb)
        
        # Conexión a la base de datos MySQL
        conn = conectar_base_de_datos()
        cursor = conn.cursor()

        cursor.execute("SELECT u.dni FROM usuario u ORDER BY u.dni ASC")
        datos = cursor.fetchall()
        self.numeros_dni = [str(item[0]) for item in datos]

        nombre = QCompleter(self.numeros_dni)
        nombre.setCaseSensitivity(Qt.CaseSensitivity.CaseInsensitive)
        nombre.setCompletionMode(QCompleter.CompletionMode.PopupCompletion)
        nombre.popup().setStyleSheet(style.completer)
        self.view_nomb.setCompleter(nombre)

        cursor.close()
        conn.close()
                        
        view_disciplina = QLabel("Disciplina:", comboView)
        view_disciplina.setStyleSheet(style.label)
        view_disciplina.setFixedWidth(100)
        self.view_disciplina = QLineEdit(comboView)
        self.view_disciplina.setPlaceholderText("Ingrese una disciplina")
        self.view_disciplina.setStyleSheet(style.estilo_lineedit)
        self.view_disciplina.setFixedWidth(200)
        elementos.addWidget(view_disciplina)       
        elementos.addWidget(self.view_disciplina)

        # Conexión a la base de datos MySQL
        conn = conectar_base_de_datos()
        cursor = conn.cursor()

        cursor.execute("SELECT d.nombre FROM disciplina d ORDER BY d.nombre ASC")
        datos = cursor.fetchall()
        self.activi = [str(item[0]) for item in datos]

        actividad = QCompleter(self.activi)
        actividad.setCaseSensitivity(Qt.CaseSensitivity.CaseInsensitive)
        actividad.setCompletionMode(QCompleter.CompletionMode.PopupCompletion)
        actividad.popup().setStyleSheet(style.completer)
        self.view_disciplina.setCompleter(actividad)

        cursor.close()
        conn.close()
             
        view_fechaDePago = QLabel("Fecha de pago:", comboView)
        view_fechaDePago.setStyleSheet(style.label)
        view_fechaDePago.setFixedWidth(140)
        self.view_fechaDePago = QDateEdit(comboView)
        self.view_fechaDePago.setStyleSheet(style.fecha)
        self.view_fechaDePago.setLocale(QLocale("es-AR"))
        self.view_fechaDePago.setCursor(Qt.CursorShape.PointingHandCursor)
        self.view_fechaDePago.setFixedWidth(200)
        self.view_fechaDePago.setDate(QDate.currentDate())
        self.view_fechaDePago.setDisplayFormat("dd/MM/yyyy")
        self.view_fechaDePago.setCalendarPopup(True)
        elementos2.addWidget(view_fechaDePago)      
        elementos2.addWidget(self.view_fechaDePago)
        elementos2.addSpacing(5)
        
        view_al2 = QLabel("Al:", comboView)
        view_al2.setStyleSheet(style.label)
        view_al2.setFixedWidth(30)
        self.view_al2 = QDateEdit(comboView)
        self.view_al2.setStyleSheet(style.fecha)
        self.view_al2.setLocale(QLocale("es-AR"))
        self.view_al2.setCursor(Qt.CursorShape.PointingHandCursor)
        self.view_al2.setFixedWidth(200)
        self.view_al2.setDate(QDate.currentDate())
        self.view_al2.setDisplayFormat("dd/MM/yyyy")
        self.view_al2.setCalendarPopup(True)
        elementos2.addWidget(view_al2)      
        elementos2.addWidget(self.view_al2)
        elementos2.setContentsMargins(0,0,50,0)     # AGREGA MARGEN ENTRE ELEMENTOS
        
        view_asistencia = QLabel("Asistentcia:", comboView)
        view_asistencia.setStyleSheet(style.label)
        view_asistencia.setFixedWidth(140)
        self.view_asistencia = QDateEdit(comboView)
        self.view_asistencia.setStyleSheet(style.fecha)
        self.view_asistencia.setLocale(QLocale("es-AR"))
        self.view_asistencia.setCursor(Qt.CursorShape.PointingHandCursor)
        self.view_asistencia.setFixedWidth(200)
        self.view_asistencia.setDate(QDate.currentDate())
        self.view_asistencia.setDisplayFormat("dd/MM/yyyy")
        self.view_asistencia.setCalendarPopup(True)
        elementos3.addWidget(view_asistencia)       
        elementos3.addWidget(self.view_asistencia)

        view_al = QLabel("Al:", comboView)
        view_al.setStyleSheet(style.label)
        view_al.setFixedWidth(30)
        self.view_al = QDateEdit(comboView)
        self.view_al.setStyleSheet(style.fecha)
        self.view_al.setLocale(QLocale("es-AR"))
        self.view_al.setCursor(Qt.CursorShape.PointingHandCursor)
        self.view_al.setFixedWidth(200)
        self.view_al.setDate(QDate.currentDate())
        self.view_al.setDisplayFormat("dd/MM/yyyy")
        self.view_al.setCalendarPopup(True)
        elementos3.addSpacing(5)
        elementos3.addWidget(view_al)       
        elementos3.addWidget(self.view_al)
        elementos3.addSpacing(10)       
        
        layout_H18 = QHBoxLayout() 
        layout_H18.setAlignment(Qt.AlignmentFlag.AlignRight)
        button14 = QPushButton('TABLA POR CLIENTE', comboView)
        button14.setFixedWidth(220)
        button14.setCursor(Qt.CursorShape.PointingHandCursor)
        button14.setStyleSheet(style.estilo_boton)
        button17 = QPushButton('TOTAL POR DISCIPLINA', comboView)
        button17.setFixedWidth(220)
        button17.setCursor(Qt.CursorShape.PointingHandCursor)
        button17.setStyleSheet(style.estilo_boton)
        button19 = QPushButton('ASISTENCIA POR CLIENTE', comboView)
        button19.setFixedWidth(220)
        button19.setCursor(Qt.CursorShape.PointingHandCursor)
        button19.setStyleSheet(style.estilo_boton)
        layout_H18.addWidget(button14)
        layout_H18.addWidget(button17)
        layout_H18.addWidget(button19)
        
        layout_H19 = QHBoxLayout()
        layout_H19.setAlignment(Qt.AlignmentFlag.AlignRight)
        button15 = QPushButton('TABLA CLIENTE', comboView)
        button15.setFixedWidth(220)
        button15.setCursor(Qt.CursorShape.PointingHandCursor)
        button15.setStyleSheet(style.estilo_boton)
        button16 = QPushButton('TOTAL DISCIPLINA', comboView)
        button16.setFixedWidth(220)
        button16.setCursor(Qt.CursorShape.PointingHandCursor)
        button16.setStyleSheet(style.estilo_boton)
        button18 = QPushButton('TOTAL ASISTENCIA', comboView)
        button18.setFixedWidth(220)
        button18.setCursor(Qt.CursorShape.PointingHandCursor)
        button18.setStyleSheet(style.estilo_boton)
        layout_H19.addWidget(button15)
        layout_H19.addWidget(button16)
        layout_H19.addWidget(button18)
        
        layout_H20 = QHBoxLayout()
        layout_H20.setAlignment(Qt.AlignmentFlag.AlignRight)
        sacar_tabla = QPushButton('LIMPIAR TABLA', comboView)
        sacar_tabla.setFixedWidth(220)
        sacar_tabla.setCursor(Qt.CursorShape.PointingHandCursor)
        sacar_tabla.setStyleSheet(style.estilo_boton)
        bottonExcel = QPushButton('DESCARGAR PLANILLA', comboView)
        bottonExcel.setFixedWidth(220)
        bottonExcel.setCursor(Qt.CursorShape.PointingHandCursor)
        bottonExcel.setStyleSheet(style.boton_excel)
        layout_H20.addWidget(sacar_tabla)
        layout_H20.addWidget(bottonExcel)
        
        # CONECCION DE SEÑALES A LAS FUNCIONES
        button14.clicked.connect(self.consultar_TotalPorAlumno)
        button15.clicked.connect(self.consultar_TotalAlumno)
        button16.clicked.connect(self.consultar_TotalDisciplina)
        sacar_tabla.clicked.connect(self.limpiar_tabla_balance)
        button17.clicked.connect(self.consultar_PorDisciplina)
        button18.clicked.connect(self.consultar_AsisteTotal)
        button19.clicked.connect(self.consultar_AsistePorAlumno)
        bottonExcel.clicked.connect(self.tabla_balance)
        
        horizontal = QHBoxLayout()
        horizontal.addLayout(elementos)
        horizontal.addLayout(layout_H18)
        horizontal2 = QHBoxLayout()
        horizontal2.addLayout(elementos2)
        horizontal2.addLayout(layout_H19)
        horizontal3 = QHBoxLayout()
        horizontal3.addLayout(elementos3)
        horizontal3.addLayout(layout_H20)

        
        # AGREGAR LAYOUT VERTICAL LOS LAYOUT HORIZONTALES
        layout_V4.addLayout(horizontal)
        layout_V4.addLayout(horizontal2)
        layout_V4.addLayout(horizontal3)
        
        # AGREGAR AL "GRID" EL LAYOUT VERTICAL
        grid5.addLayout(layout_V4,0,0,1,5)
        
        # CREA LA TABLA DENTRO DEL COMBOBOX
        self.tablaVIEW = QTableWidget(comboView)
        self.tablaVIEW.setCursor(Qt.CursorShape.PointingHandCursor)
        self.tablaVIEW.setStyleSheet(style.esttabla)
        
        # AGREGAR AL "GRID" LA TABLA
        grid5.addWidget(self.tablaVIEW,1,0,1,5)
        
        # AGREAGA EL QCOMBOBOX AL GRID
        comboView.setLayout(grid5)
        
        # Agregar el QGroupBox a la primera pestaña (tab1)
        tab5_layout = QVBoxLayout()
        tab5_layout.addWidget(comboView)
        pestania_view.setLayout(tab5_layout)
    
        #-----------------------------------------------------------------
        # PESTAÑA REGISTRAR DE LOS EMPLEADOS
        grupo_empleados = QGroupBox("DETALLE DEL EMPLEADO", pestania_empleados)
        grupo_empleados.setStyleSheet(style.estilo_grupo)
        
        # Colocar el ComboBox a la grilla
        grid_emp = QGridLayout(grupo_empleados)
        
        # CONTENEDOR DE LOS LAYOUT HORIZONTALES
        vertical_v = QVBoxLayout()
 
        # Crear un diseño para elementos '''LABEL y QLineEdit''' al diseño del QGroupBoxel QGroupBox
        layout_emp = QHBoxLayout()
        layout_emp.setAlignment(Qt.AlignmentFlag.AlignLeft)
        layout_emp1 = QHBoxLayout()
        layout_emp1.setAlignment(Qt.AlignmentFlag.AlignLeft)
        
        # Definir el atributo id_empleado_seleccionado
        self.id_empleado_seleccionado = None
        
        id_horas_empleado = QLabel('Nombre:',grupo_empleados)
        id_horas_empleado.setStyleSheet(style.label)
        id_horas_empleado.setFixedWidth(80)
        self.id_horas_empleado = QLineEdit(grupo_empleados)
        self.id_horas_empleado.setStyleSheet(style.estilo_lineedit)
        self.id_horas_empleado.setFixedWidth(200)
        self.id_horas_empleado.setPlaceholderText("Nombre del empleado")
        self.id_horas_empleado.setFocus()
        layout_emp.addWidget(id_horas_empleado)        
        layout_emp.addWidget(self.id_horas_empleado)
        
        # Conexión a la base de datos MySQL
        conn = conectar_base_de_datos()
        cursor = conn.cursor()

        # Consulta para obtener los datos de una columna específica
        cursor.execute("SELECT id_empleado, nombre FROM registro_empleado WHERE habilitado = 1 ORDER BY nombre ASC")
        datos = cursor.fetchall()
        self.listas = [(str(item[1]), item[0]) for item in datos]  # Crear una lista de tuplas (nombre, id_empleado)
        
        lista_empleado = QCompleter([item[0] for item in self.listas])  # Usar solo los nombres para la lista desplegable
        lista_empleado.setCaseSensitivity(Qt.CaseSensitivity.CaseInsensitive)
        lista_empleado.setCompletionMode(QCompleter.CompletionMode.PopupCompletion)
        lista_empleado.popup().setStyleSheet(style.completer)
        self.id_horas_empleado.setCompleter(lista_empleado)
        
        # Agregar evento de selección de elemento del completer
        lista_empleado.activated[str].connect(self.guardar_id_empleado)#[str]
        
        cursor.close()
        conn.close()
        
        horas_tra = QLabel('Horas diarias:',grupo_empleados)
        horas_tra.setStyleSheet(style.label)
        horas_tra.setFixedWidth(120)
        self.horas_tra = QLineEdit(grupo_empleados)
        self.horas_tra.setStyleSheet(style.estilo_lineedit)
        self.horas_tra.setFixedWidth(100)
        layout_emp.addWidget(horas_tra)      
        layout_emp.addWidget(self.horas_tra)
        
        fecha_tra = QLabel('Fecha:',grupo_empleados)
        fecha_tra.setStyleSheet(style.label)
        fecha_tra.setFixedWidth(60)
        self.fecha_tra = QDateEdit(grupo_empleados)
        self.fecha_tra.setLocale(QLocale("es-AR"))
        self.fecha_tra.setStyleSheet(style.fecha)
        self.fecha_tra.setCursor(Qt.CursorShape.PointingHandCursor)
        self.fecha_tra.setFixedWidth(200)
        self.fecha_tra.setDate(QDate.currentDate())
        self.fecha_tra.setDisplayFormat("dd/MM/yyyy")
        self.fecha_tra.setCalendarPopup(True)
        layout_emp.addWidget(fecha_tra)      
        layout_emp.addWidget(self.fecha_tra)
        
        periodo = QLabel("Período:", grupo_empleados)
        periodo.setStyleSheet(style.label)
        periodo.setFixedWidth(80)
        self.periodo = QDateEdit(grupo_empleados)
        self.periodo.setStyleSheet(style.fecha)
        self.periodo.setLocale(QLocale("es-AR"))
        self.periodo.setCursor(Qt.CursorShape.PointingHandCursor)
        self.periodo.setFixedWidth(200)
        self.periodo.setDate(QDate.currentDate())
        self.periodo.setDisplayFormat("dd/MM/yyyy")
        self.periodo.setCalendarPopup(True)
        layout_emp1.addWidget(periodo)       
        layout_emp1.addWidget(self.periodo)
        
        fin_tra = QLabel('Al:',grupo_empleados)
        fin_tra.setStyleSheet(style.label)
        fin_tra.setFixedWidth(50)
        self.fin_tra = QDateEdit(grupo_empleados)
        self.fin_tra.setLocale(QLocale("es-AR"))
        self.fin_tra.setStyleSheet(style.fecha)
        self.fin_tra.setCursor(Qt.CursorShape.PointingHandCursor)
        self.fin_tra.setFixedWidth(200)
        self.fin_tra.setDate(QDate.currentDate())
        self.fin_tra.setDisplayFormat("dd/MM/yyyy")
        self.fin_tra.setCalendarPopup(True)
        layout_emp1.addWidget(fin_tra)       
        layout_emp1.addWidget(self.fin_tra)
        
        # LAYOUT HORIZONTAL PARA BOTONES
        emp3 = QHBoxLayout()
        emp3.setAlignment(Qt.AlignmentFlag.AlignRight)
        guardar_hora = QPushButton('GUARDAR', grupo_empleados)
        guardar_hora.setFixedWidth(200)
        guardar_hora.setCursor(Qt.CursorShape.PointingHandCursor)
        guardar_hora.setStyleSheet(style.estilo_boton)
        guardar_hora.clicked.connect(self.guardar_horas)
        
        button02 = QPushButton('PERIODO DE HORAS', grupo_empleados)
        button02.setFixedWidth(200)
        button02.setCursor(Qt.CursorShape.PointingHandCursor)
        button02.setStyleSheet(style.estilo_boton)
        button02.clicked.connect(self.horas_empleado_totales)
        
        emp3.addWidget(guardar_hora)
        emp3.addWidget(button02)
        
        emp4 =  QHBoxLayout()
        emp4.setAlignment(Qt.AlignmentFlag.AlignRight)
        actualizar_hoas = QPushButton('ACTUALIZAR', grupo_empleados)
        actualizar_hoas.setCursor(Qt.CursorShape.PointingHandCursor)
        actualizar_hoas.setFixedWidth(200)
        actualizar_hoas.setStyleSheet(style.estilo_boton)
        actualizar_hoas.clicked.connect(self.actualizar_horas)
        
        eliminar_horas = QPushButton('ELIMINAR', grupo_empleados)
        eliminar_horas.setFixedWidth(200)
        eliminar_horas.setCursor(Qt.CursorShape.PointingHandCursor)
        eliminar_horas.setStyleSheet(style.estilo_boton)
        eliminar_horas.clicked.connect(self.eliminar_horas)
        
        emp4.addWidget(actualizar_hoas)
        emp4.addWidget(eliminar_horas)
                
        primer = QHBoxLayout()
        primer.addLayout(layout_emp)
        primer.addLayout(emp3)
        segundo = QHBoxLayout()
        segundo.addLayout(layout_emp1)
        segundo.addLayout(emp4)
        
        # AGREDA LAYOUT HORIZONTALES AL LAYOUT VERTICAL
        vertical_v.addLayout(primer)
        vertical_v.addLayout(segundo)
        
        grid_emp.addLayout(vertical_v,0,0,1,5)
        
        # CREA LA TABLA
        self.tablaHoras = QTableWidget(grupo_empleados)
        self.tablaHoras.setCursor(Qt.CursorShape.PointingHandCursor)
        self.tablaHoras.setStyleSheet(style.esttabla)
        self.tablaHoras.clicked.connect(self.autocompleto_de_datos_horas)
        self.tablaHoras.itemSelectionChanged.connect(self.seleccionar_fila)
        
        self.empleado = QPushButton('EMPLEADO', grupo_empleados)
        self.empleado.setStyleSheet(style.estilo_boton)
        self.empleado.setCursor(Qt.CursorShape.PointingHandCursor)
        self.empleado.clicked.connect(self.emp)
        
        limpiarTABLA = QPushButton('LIMPIAR TABLA', grupo_empleados)
        limpiarTABLA.setFixedWidth(200)
        limpiarTABLA.setCursor(Qt.CursorShape.PointingHandCursor)
        limpiarTABLA.setStyleSheet(style.estilo_boton)
        limpiarTABLA.clicked.connect(self.limpiar_tabla_horas)
        
        button_horas_por_empleado = QPushButton('HORAS POR EMPLEADO', grupo_empleados)
        button_horas_por_empleado.setFixedWidth(200)
        button_horas_por_empleado.setCursor(Qt.CursorShape.PointingHandCursor)
        button_horas_por_empleado.setStyleSheet(style.estilo_boton)
        button_horas_por_empleado.clicked.connect(self.horas_empleado)
        
        execel_horas = QPushButton('DESCARGAR PLANILLA', grupo_empleados)
        execel_horas.setFixedWidth(200)
        execel_horas.setCursor(Qt.CursorShape.PointingHandCursor)
        execel_horas.setStyleSheet(style.boton_excel)
        execel_horas.clicked.connect(self.excel_horas)
        
        # AGREDA LA TABLA y BOTON A LA GRILLA 
        h = QHBoxLayout()
        
        v = QVBoxLayout()
        v.setAlignment(Qt.AlignmentFlag.AlignTop)
        v.addSpacing(20)
        v.addWidget(self.empleado)
        v.addSpacing(10)
        v.addWidget(limpiarTABLA)
        v.addSpacing(10)
        v.addWidget(button_horas_por_empleado)
        v.addSpacing(10)
        v.addWidget(execel_horas)
        
        h.addWidget(self.tablaHoras)
        h.addSpacing(25)
        h.addLayout(v)
        h.addSpacing(25)
        grid_emp.addLayout(h,1,0,1,5)       
        
        # Establecer el diseño del QGroupBox
        grupo_empleados.setLayout(grid_emp)
        
        # Agregar el QGroupBox a la primera pestaña (tab1)
        tab_emp_layout = QVBoxLayout()
        tab_emp_layout.addWidget(grupo_empleados)
        pestania_empleados.setLayout(tab_emp_layout)
                
        #-----------------------------------------------------------------
        # PESTAÑA REGISTRAR GASTOS
        
        # SE CREA ComboBox 'REGISTRAR GASTOS'
        grupo_resumen = QGroupBox("DETALLE CONTABLE", pestania_resumen)
        grupo_resumen.setStyleSheet(style.estilo_grupo)
        # Colocar el ComboBox a la grilla
        grid_resumen = QGridLayout(grupo_resumen)
        
        # LAYOUT VERTICAL PARA LOS LAYOUT HORIZONTAL
        vertical = QVBoxLayout()
        
        # LAYOUT HORIZONTAL PARA LOS ELEMENTOS
        layout_libro = QHBoxLayout()
        layout_libro.setAlignment(Qt.AlignmentFlag.AlignLeft)
      
        fecha_gastos = QLabel('Fecha:',grupo_resumen)
        fecha_gastos.setStyleSheet(style.label)
        fecha_gastos.setFixedWidth(140)
        self.fecha_gastos = QDateEdit(grupo_resumen)
        self.fecha_gastos.setLocale(QLocale("es-AR"))
        self.fecha_gastos.setStyleSheet(style.fecha)
        self.fecha_gastos.setCursor(Qt.CursorShape.PointingHandCursor)
        self.fecha_gastos.setFixedWidth(200)
        self.fecha_gastos.setCalendarPopup(True)
        self.fecha_gastos.setDisplayFormat("dd/MM/yyyy")
        self.fecha_gastos.setDate(QDate.currentDate())
        layout_libro.addWidget(fecha_gastos)  
        layout_libro.addWidget(self.fecha_gastos)
        
        # LAYOUT HORIZONTAL PARA LOS ELEMENTOS
        layout_conepto = QHBoxLayout()
        layout_conepto.setAlignment(Qt.AlignmentFlag.AlignLeft)
                
        concepto_debe = QLabel('Cuenta (Debe):',grupo_resumen)
        concepto_debe.setStyleSheet(style.label)
        concepto_debe.setFixedWidth(140)
        self.concepto_debe = QLineEdit(grupo_resumen)
        self.concepto_debe.textChanged.connect(self.debe)
        self.concepto_debe.setStyleSheet(style.estilo_lineedit)
        self.concepto_debe.setFixedWidth(200)
        self.concepto_debe.setFocus()
        layout_conepto.addWidget(concepto_debe)  
        layout_conepto.addWidget(self.concepto_debe)
        layout_conepto.addSpacing(5)
        
        # Conexión a la base de datos MySQL
        conn = conectar_base_de_datos()
        cursor = conn.cursor()

        # Consulta para obtener los datos de una columna específica
        cursor.execute("SELECT nombre FROM cuenta WHERE categoria = 'debe'")
        dato = cursor.fetchall()
        self.sugerencia1 = [str(item[0]) for item in dato]
        print(self.sugerencia1)

        lista_debe = QCompleter(self.sugerencia1)
        lista_debe.setCaseSensitivity(Qt.CaseSensitivity.CaseInsensitive)
        lista_debe.setCompletionMode(QCompleter.CompletionMode.PopupCompletion)
        lista_debe.popup().setStyleSheet(style.completer)
        self.concepto_debe.setCompleter(lista_debe)
        
        cursor.close()
        conn.close()
                
        concepto_haber = QLabel('Cuenta (Haber):',grupo_resumen)
        concepto_haber.setStyleSheet(style.label)
        concepto_haber.setFixedWidth(140)
        self.concepto_haber = QLineEdit(grupo_resumen)
        self.concepto_haber.textChanged.connect(self.haber)
        self.concepto_haber.setStyleSheet(style.estilo_lineedit)
        self.concepto_haber.setFixedWidth(200)
        layout_conepto.addWidget(concepto_haber)
        layout_conepto.addWidget(self.concepto_haber)
        layout_conepto.addSpacing(5)
        
        # Conexión a la base de datos MySQL
        conn = conectar_base_de_datos()
        cursor = conn.cursor()

        # Consulta para obtener los datos de una columna específica
        cursor.execute("SELECT nombre FROM cuenta WHERE categoria = 'haber'")
        datos = cursor.fetchall()
        self.sugerencia2 = [str(item[0]) for item in datos]
        print(self.sugerencia2)
        
        lista_haber = QCompleter(self.sugerencia2)
        lista_haber.setCaseSensitivity(Qt.CaseSensitivity.CaseInsensitive)
        lista_haber.setCompletionMode(QCompleter.CompletionMode.PopupCompletion)
        lista_haber.popup().setStyleSheet(style.completer)
        self.concepto_haber.setCompleter(lista_haber)
        
        cursor.close()
        conn.close()
               
        debe = QLabel('Debe ($):',grupo_resumen)
        debe.setStyleSheet(style.label)
        debe.setFixedWidth(80)
        self.debe = QLineEdit(grupo_resumen)
        self.debe.setStyleSheet(style.estilo_lineedit)
        self.debe.setFixedWidth(100)
        layout_conepto.addWidget(debe)
        layout_conepto.addWidget(self.debe)
        layout_conepto.addSpacing(5) 
        
        haber = QLabel('Haber ($):',grupo_resumen)
        haber.setStyleSheet(style.label)
        haber.setFixedWidth(90)
        self.haber = QLineEdit(grupo_resumen)
        self.haber.setStyleSheet(style.estilo_lineedit)
        self.haber.setFixedWidth(100)
        layout_conepto.addWidget(haber)
        layout_conepto.addWidget(self.haber)
        
        layout_libro3 = QHBoxLayout()
        layout_libro3.setAlignment(Qt.AlignmentFlag.AlignLeft)
        fecha_periodo = QLabel('Periodo:',grupo_resumen)
        fecha_periodo.setStyleSheet(style.label)
        fecha_periodo.setFixedWidth(140)
        self.fecha_periodo = QDateEdit(grupo_resumen)
        self.fecha_periodo.setStyleSheet(style.fecha)
        self.fecha_periodo.setLocale(QLocale("es-AR"))
        self.fecha_periodo.setCursor(Qt.CursorShape.PointingHandCursor)
        self.fecha_periodo.setFixedWidth(200)
        self.fecha_periodo.setCalendarPopup(True)
        self.fecha_periodo.setDisplayFormat("dd/MM/yyyy")
        self.fecha_periodo.setDate(QDate.currentDate())
        layout_libro3.addWidget(fecha_periodo)   
        layout_libro3.addWidget(self.fecha_periodo)
        layout_libro3.addSpacing(5) # AGREGA ESPACIO ENTRE ELEMENTOS
        
        fecha_fin = QLabel('Al:',grupo_resumen)
        fecha_fin.setStyleSheet(style.label)
        fecha_fin.setFixedWidth(45)
        self.fecha_fin = QDateEdit(grupo_resumen)
        self.fecha_fin.setStyleSheet(style.fecha)
        self.fecha_fin.setLocale(QLocale("es-AR"))
        self.fecha_fin.setCursor(Qt.CursorShape.PointingHandCursor)
        self.fecha_fin.setFixedWidth(200)
        self.fecha_fin.setCalendarPopup(True)
        self.fecha_fin.setDisplayFormat("dd/MM/yyyy")
        self.fecha_fin.setDate(QDate.currentDate())
        layout_libro3.addWidget(fecha_fin)   
        layout_libro3.addWidget(self.fecha_fin)
        
        # AGREDA LOS LAYOUT HORIZONTALES AL LAYOUT VERTICAL
        vertical.addLayout(layout_libro)
        vertical.addLayout(layout_conepto)
        vertical.addLayout(layout_libro3)
                
        # AGREDA LOS LAYOUT VERTICAL A LA GRILLA
        grid_resumen.addLayout(vertical,0,0,1,5)
        
        # CREA UNA TABLA QUE VA A ESTAR DENTRO DEL QGROUPBOX
        self.tablaGastos = QTableWidget(grupo_resumen)
        self.tablaGastos.setCursor(Qt.CursorShape.PointingHandCursor)
        self.tablaGastos.setStyleSheet(style.esttabla)
        self.tablaGastos.clicked.connect(self.selecionarTabla)
        self.tablaGastos.itemSelectionChanged.connect(self.seleccionar_fila)
        
        buttonREG = QPushButton('REGISTRAR', grupo_resumen)
        buttonREG.setFixedWidth(200)
        buttonREG.setCursor(Qt.CursorShape.PointingHandCursor)
        buttonREG.setStyleSheet(style.estilo_boton)
        buttonREG.clicked.connect(self.registrar_datos)
        
        buttonACT = QPushButton('ACTUALIZAR', grupo_resumen)
        buttonACT.setFixedWidth(200)
        buttonACT.setCursor(Qt.CursorShape.PointingHandCursor)
        buttonACT.setStyleSheet(style.estilo_boton)
        buttonACT.clicked.connect(self.actualizar_datos)
        
        buttonPERIODO = QPushButton('VISUALIZAR PERIODO', grupo_resumen)
        buttonPERIODO.setFixedWidth(200)
        buttonPERIODO.setCursor(Qt.CursorShape.PointingHandCursor)
        buttonPERIODO.setStyleSheet(style.estilo_boton)
        buttonPERIODO.clicked.connect(self.visualizacion_datos)
        
        button_eliminar = QPushButton('ELIMINAR', grupo_resumen)
        button_eliminar.setFixedWidth(200)
        button_eliminar.setCursor(Qt.CursorShape.PointingHandCursor)
        button_eliminar.setStyleSheet(style.estilo_boton)
        button_eliminar.clicked.connect(self.eliminar_datos)
        
        cuenta = QPushButton('CARGAR CUENTA', grupo_resumen)
        cuenta.setFixedWidth(200)
        cuenta.setCursor(Qt.CursorShape.PointingHandCursor)
        cuenta.setStyleSheet(style.estilo_boton)
        cuenta.clicked.connect(self.cargar_cuenta)
        
        boton_limpiarTabla = QPushButton('LIMPIAR TABLA', grupo_resumen)
        boton_limpiarTabla.setFixedWidth(200)
        boton_limpiarTabla.setStyleSheet(style.estilo_boton)
        boton_limpiarTabla.setCursor(Qt.CursorShape.PointingHandCursor)
        boton_limpiarTabla.clicked.connect(self.limp_tabla)
        
        excel_resumen = QPushButton('DESCARGAR PLANILLA', grupo_resumen)
        excel_resumen.setFixedWidth(200)
        excel_resumen.setCursor(Qt.CursorShape.PointingHandCursor)
        excel_resumen.setStyleSheet(style.boton_excel)
        excel_resumen.clicked.connect(self.tabla_resumen)  
        
        # AGREDA LA TABLA y BOTON A LA GRILLA 
        costado = QVBoxLayout()
        costado.setAlignment(Qt.AlignmentFlag.AlignTop)
        costado.addSpacing(10)
        costado.addWidget(buttonREG)
        costado.addSpacing(10)
        costado.addWidget(buttonPERIODO)
        costado.addSpacing(10)
        costado.addWidget(buttonACT)
        costado.addSpacing(10)
        costado.addWidget(button_eliminar)
        costado.addSpacing(30)
        costado.addWidget(cuenta)
        costado.addSpacing(10)
        costado.addWidget(boton_limpiarTabla)
        costado.addSpacing(10)
        costado.addWidget(excel_resumen)
        
        hori2 = QHBoxLayout()
        hori2.addWidget(self.tablaGastos)
        hori2.addSpacing(25)
        hori2.addLayout(costado)
        hori2.addSpacing(25)
        grid_resumen.addLayout(hori2,1,0,1,5)
    
        # Establecer el diseño del QGroupBox
        grupo_resumen.setLayout(grid_resumen)
        
        # Agregar el QGroupBox a la primera pestaña (tab1)
        tab_gastos_layout = QVBoxLayout()
        tab_gastos_layout.addWidget(grupo_resumen)
        pestania_resumen.setLayout(tab_gastos_layout)
        #-----------------------------------------------------------------
        
        # Crear un layout horizontal para contener layout_vertical1 y widget_contenedor2
        layout_horizontal = QHBoxLayout()
        layout_horizontal.addLayout(layout_vertical1) # FRAME
        layout_horizontal.addWidget(self.tab) # TAB
        layout_horizontal.setContentsMargins(0, 0, 0, 0)
        
        # Establecer el layout horizontal como el layout de la ventana principal
        central_widget = QFrame()
        central_widget.setStyleSheet("background-color: black;")
        central_widget.setLayout(layout_horizontal)
        
        # AGREGA AL FRAME(MARCO) AL WIDGET CENTRAL
        self.setCentralWidget(central_widget)
        
        # BLOQUEO DE BOTON ELIMINAR
        if not self.tableActivi.clicked:
            button_Elim.setEnabled(False)
        else:
            button_Elim.setEnabled(True)
            
    def acciones(self):
        # Acción para crear un archivo de texto
        self.create_text_file_action = QAction('&Crear archivo de texto', self)
        self.create_text_file_action.setShortcut(QKeySequence('Ctrl+N'))
        self.create_text_file_action.setStatusTip('Crear un nuevo archivo de texto')
        self.create_text_file_action.triggered.connect(self.createTextFile)
        
        self.exit_action = QAction('&Cerrar', self)
        self.exit_action.setShortcut(QKeySequence('Ctrl+Q'))
        self.exit_action.setStatusTip('Salir de la aplicación')
        self.exit_action.triggered.connect(self.close)
        
        # Acción separadora
        self.separator_action = QAction(self)
        self.separator_action.setSeparator(True)
        self.separator_action.setVisible(True)
        
        menubar = self.menuBar()
        menubar.setStyleSheet(style.estilo_menubar)
        file_menu = menubar.addMenu('&Archivo')
        file_menu.addAction(self.create_text_file_action)
        file_menu.addAction(self.separator_action)
        file_menu.addAction(self.exit_action)
    
    def createTextFile(self):
        
        self.texto = ArchivoTexto()
        self.texto.show()
        
    # FUNCIONES PARA VINCULAR EL QTabWidget
    def record(self):
        self.tab.setCurrentIndex(0)
        self.tab.setDisabled(False)
        self.input_nombre1.setFocus()
        
    def update(self):
        self.tab.setCurrentIndex(1)
        self.tab.setDisabled(False)
        self.input_nombre2.setFocus()

    def deleteRecord(self):
        self.tab.setCurrentIndex(2)
        self.tab.setDisabled(False)
        self.nombre_buscar3.setFocus()
        
    def activity(self):
        self.tab.setCurrentIndex(3)
        self.tab.setDisabled(False)
        self.input_disciplina4.setFocus()
        
    def pagos(self):
        self.tab.setCurrentIndex(4)
        self.tab.setDisabled(False)    
        self.idUser.setFocus()
    
    def empleados(self):
        self.tab.setCurrentIndex(5)
        self.tab.setDisabled(False)
    
    def emp(self): # REGISTRO DE EMPLEADO
        self.cargaEmple = Empleado()
        self.cargaEmple.show()
        
    def balances(self):
        self.tab.setCurrentIndex(6)
        self.tab.setDisabled(False)
        
    def registro_de_ingYegreso(self):
        self.tab.setCurrentIndex(7)
        self.tab.setDisabled(False)
        self.concepto_debe.setFocus()
    
    def cargar_cuenta(self):
        self.tipo_cuenta = CuentaContable()
        self.tipo_cuenta.show()
        
    # FUNCION QUE VINCULA LA VENTANA DE ASISTENCIA
    def assistance(self):
        self.boton = Asistencia()
        self.boton.show()
    
    # Pestaña de REGISTRO ---------------------------------------
    def guardar(self):
        nombre1 = self.input_nombre1.text().title()
        apellido1 = self.input_apellido1.text().title()
        dni = self.input_dni.text().replace(".", "")
        sexo = self.input_sex.currentData()
        edad = self.input_age.text()
        celu = self.input_celular.text().replace(".", "")
        fecha = self.input_date.date().toPyDate()

        validacion = registroUSER(self, nombre1, apellido1, dni, sexo, edad, celu)
        if validacion != "Validación exitosa.":
            mensaje_ingreso_datos("Error de validación", "Verifique los datos por favor")
            return      
        
        try:
            db = conectar_base_de_datos()
            cursor = db.cursor()

            # Verificar si el dni ya existe
            cursor.execute("SELECT COUNT(*) FROM usuario WHERE dni = %s", (dni,))
            resultado = cursor.fetchone()

            if resultado[0] > 0:
                # Si el dni ya existe, mostrar un mensaje y no insertar
                ingreso_datos("Registro de cliente", "El DNI ya está registrado")
                limpiasElementosUser(self,QDate)
            else:
                # Si el dni no existe, insertar el nuevo usuario
                query_insertar = "INSERT INTO usuario (nombre, apellido, dni, sexo, edad, celular, fecha_registro) VALUES (%s, %s, %s, %s, %s, %s, %s)"
                values = (nombre1, apellido1, dni, sexo, edad, celu, fecha)
                cursor.execute(query_insertar, values)
                db.commit()

                if cursor.rowcount > 0:
                    ingreso_datos("Registro de cliente", "Registro cargado")
                    limpiasElementosUser(self, QDate)
                    self.onChangeTab(0)
                else:
                    ingreso_datos("Registro de cliente", "Registro no cargado")

            cursor.close()
            db.close()
        except Error as ex:
            errorConsulta("Registro de cliente", f"Error en la consulta: {str(ex)}")
    
    def search(self):
        if not self.input_nombre1.text():
            mensaje_ingreso_datos("Registro de cliente","Introduzca el nombre del alumno a buscar")
            return
        
        nombre = self.input_nombre1.text()       
        
        patron = re.compile(r'^[a-zA-ZáéíóúÁÉÍÓÚüÜ\'\s]+$') 
        if not isinstance(nombre, str) or nombre.isspace() or not patron.match(nombre): 
            mensaje_ingreso_datos("Registro de cliente","El nombre debe contener: \n- Letras y/o espacios entre nombres(si tiene mas de dos).")
            return
        
        try:
            db = conectar_base_de_datos()
            cursor = db.cursor()
            cursor.execute(f"SELECT id_usuario AS ID, nombre, apellido, dni, sexo, edad, celular, fecha_registro AS REGISTRO FROM usuario WHERE habilitado = 1 AND nombre LIKE '%{nombre}%' ORDER BY nombre ASC")
            # cursor.execute(f"SELECT * FROM usuario WHERE nombre LIKE '%{nombre}%'")
            resultados = cursor.fetchall()
            
            if len(resultados) > 0:
                ingreso_datos("Registro de cliente",f"Se encontraron {len(resultados)} coincidencias.")
                self.input_nombre1.clear()
                tabla_registroUSER(self, cursor, resultados, QHeaderView, QTableWidget, QAbstractItemView, QTableWidgetItem, QDate, Qt)
                
                self.tablaRecord.clearSelection()  # Deseleccionar la fila eliminada
            else:
                ingreso_datos("Registro de cliente",f"Se encontraron {len(resultados)} coincidencias.")
                
            # cierre de la BD
            db.close()
            cursor.close()
            
        except Error as ex:
            errorConsulta("Registro de clientes",f"Error en la consulta: {str(ex)}")
            print("Error executing the query", ex) 
        
    def mostrarTabla(self): 
        try:
            db = conectar_base_de_datos()
            cursor = db.cursor()
            cursor.execute("SELECT id_usuario AS ID, nombre, apellido, dni, sexo, edad, celular, fecha_registro AS REGISTRO FROM usuario WHERE habilitado = 1 ORDER BY nombre ASC")
            resultados = cursor.fetchall()
                    
            if len(resultados) > 0:
                tabla_registroUSER(self, cursor, resultados, QHeaderView, QTableWidget, QAbstractItemView, QTableWidgetItem, QDate, Qt)
                               
                self.tablaRecord.clearSelection()  # Deseleccionar la fila eliminada
            else:
                ingreso_datos("Registro de cliente","Tabla sin registros")
                
            cursor.close()
            db.close()
            
        except Error as ex:
            errorConsulta("Registro de cliente",f"Error en la consulta: {str(ex)}")
            
    def claer_tabla(self):
        limpiar_tablaRecord(self)
               
    def tabla_registro(self):
        tabla_registroUSUARIO(self,Workbook,Font,PatternFill,Border,Side,numbers,QFileDialog)

    # Pestaña de ACTUALIZAR REGISTRO ---------------------------------------
    def ver(self):
        self.tablaUpdateRecord.setEnabled(True)
        try:
            db = conectar_base_de_datos()
            cursor = db.cursor()
            cursor.execute("SELECT id_usuario AS ID, nombre, apellido, dni, sexo, edad, celular, fecha_registro AS REGISTRO FROM usuario WHERE habilitado = 1 ORDER BY nombre ASC")
            resultados = cursor.fetchall()

            if len(resultados) > 0:
                tabla_updateUSER(self, cursor, resultados, QHeaderView, QTableWidget, QAbstractItemView, QTableWidgetItem, QDate, Qt)
                
                self.tablaUpdateRecord.clearSelection()  # Deseleccionar la fila eliminada
            else:
                ingreso_datos("Registro de cliente","Tabla sin registros")
                    
            cursor.close()  
            db.close()
            
        except Error as ex:
            errorConsulta("Registro de cliente",f"Error en la consulta: {str(ex)}")
            print("Error executing the query", ex) 
     
    def buscar(self):
        self.tablaUpdateRecord.setEnabled(False)
        if not self.input_nombre2.text():
            mensaje_ingreso_datos("Registro de cliente","Introduzca el nombre del cliente a buscar")
            return
        else:
            self.tablaUpdateRecord.setEnabled(True)
        
        nombre_seleccionado = self.input_nombre2.text() 
        patron_nom = re.compile(r'^[a-zA-ZáéíóúÁÉÍÓÚüÜ\'\s]+$') 
        if not isinstance(nombre_seleccionado, str) or nombre_seleccionado.isspace() or not patron_nom.match(nombre_seleccionado): 
            mensaje_ingreso_datos("Registro de cliente","El nombre debe contener: \n- Letras y/o espacios entre nombres(si tiene mas de dos).")
            return      
        
        try:
            db = conectar_base_de_datos()
            cursor = db.cursor()
            # cursor.execute("SELECT * FROM usuario ORDER BY nombre ASC")
            cursor.execute(f"SELECT id_usuario AS ID, nombre, apellido, dni, sexo, edad, celular, fecha_registro AS REGISTRO FROM usuario WHERE habilitado = 1 AND nombre LIKE '%{nombre_seleccionado}%' ORDER BY nombre ASC")
            resultados = cursor.fetchall()
            
            if len(resultados) > 0:
                ingreso_datos("Registro de cliente",f"Se encontraron {len(resultados)} coincidencias.")
                self.input_nombre2.clear()
                tabla_updateUSER(self, cursor, resultados, QHeaderView, QTableWidget, QAbstractItemView, QTableWidgetItem, QDate, Qt)
                self.tablaUpdateRecord.clearSelection()  # Deseleccionar la fila eliminada
            else:
                ingreso_datos("Registro de cliente",f"Se encontraron {len(resultados)} coincidencias.")
                self.tablaUpdateRecord.clearSelection()  # Deseleccionar la fila eliminada
                
            cursor.close()
            db.close()
            
        except Error as ex:
            errorConsulta("Registro de cliente",f"Error en la consulta: {str(ex)}")
            print("Error executing the query", ex)
    
    def limpiar(self,):
        limpiar_campos(self)
    
    def seleccionYautoCompletado(self):
        autoCompletadoACTULIZAR(self,QDate)
        
    def actualizar(self):
        if not self.tablaUpdateRecord.currentItem():
            mensaje_ingreso_datos("Registro de cliente","Por favor seleccione un registro para actualizar")
            return
        
        id_reg = int(self.tablaUpdateRecord.item(self.tablaUpdateRecord.currentRow(), 0).text())
        nombre2 = self.input_nombre2.text().title()
        apellido2 = self.input_apellido2.text().title()
        dni2 = self.input_dni2.text()
        sexo2 = self.input_sex2.currentData()
        edad2 = self.input_age2.text()
        celu2 = self.input_celular2.text()
        fecha = self.input_date2.date().toPyDate()
        
        actualizacionUSER(self, nombre2 , apellido2, dni2, sexo2, edad2, celu2)
                    
        # Validar que los campos requeridos no estén vacíos
        if not nombre2 or not apellido2 or not dni2 or not sexo2 or not edad2 or not celu2:
            mensaje_ingreso_datos("Registro de cliente", "Todos los campos son obligatorios")
            return
        
        try:
            db = conectar_base_de_datos()
            cursor = db.cursor()                
            query = "UPDATE usuario SET nombre=%s, apellido=%s, dni=%s, sexo=%s, edad=%s, celular=%s, fecha_registro=%s WHERE id_usuario=%s ORDER BY nombre ASC"
            values = (nombre2, apellido2, dni2, sexo2, edad2, celu2, fecha, id_reg) 
            cursor.execute(query, values)
            db.commit()

            if cursor.rowcount > 0:
                ingreso_datos("Registro de cliente","Registro actualizado")
                limpiasElementosUseraActualizar(self,QDate)
                limpiar_tablaUpdate(self)
                self.onChangeTab(1)
                self.tablaUpdateRecord.clearSelection()  # Deseleccionar la fila eliminada
            else:
                ingreso_datos("Registro de cliente","Registro no actualizado")
                                    
            cursor.close()
            db.close()
                            
        except Error as ex:
            errorConsulta("Registro de cliente",f"Error en la consulta: {str(ex)}")
            print("Error al ejecutar la consulta", ex)      
        
    # Pestaña de ELIMINAR ---------------------------------
    def buscar_para_eliminar(self):
        self.tablaDeleteRecord.setEnabled(False)
        if not self.nombre_buscar3.text():
            mensaje_ingreso_datos("Registro de cliente","Introduzca el nombre del alumno a buscar")
            return
        
        self.tablaDeleteRecord.setEnabled(True)
            
        nombre_seleccionado3 = self.nombre_buscar3.text()
        patron_nom2 = re.compile(r'^[a-zA-ZáéíóúÁÉÍÓÚüÜ\'\s]+$') 
        if not isinstance(nombre_seleccionado3, str) or nombre_seleccionado3.isspace() or not patron_nom2.match(nombre_seleccionado3): 
            mensaje_ingreso_datos("Registro de cliente","El nombre debe contener: \n- Letras y/o espacios entre nombres(si tiene mas de dos).")
            return 
               
        try:
            db = conectar_base_de_datos()
            cursor = db.cursor()
            cursor.execute(f"SELECT id_usuario AS ID, nombre, apellido, dni, sexo, edad, celular, fecha_registro AS REGISTRO FROM usuario WHERE habilitado = 1 AND nombre LIKE '%{nombre_seleccionado3}%' ORDER BY nombre ASC")
            # cursor.execute(f"SELECT * FROM usuario WHERE nombre LIKE '%{nombre_seleccionado3}%'")
            resultados = cursor.fetchall()
            
            if len(resultados) > 0:
                ingreso_datos("Registro de cliente",f"Se encontraron {len(resultados)} coincidencias.")
                
                self.nombre_buscar3.clear()
                tabla_eliminarUSER(self, cursor, resultados, QHeaderView, QTableWidget, QAbstractItemView, QTableWidgetItem, QDate, Qt)

                self.tablaDeleteRecord.clearSelection()  # Deseleccionar la fila eliminada
            else:
                ingreso_datos("Registro de cliente",f"Se encontraron {len(resultados)} coincidencias.")
            
            cursor.close()
            db.close()
            
        except Error as ex:
            errorConsulta("Registro de cliente",f"Error en la consulta: {str(ex)}")
            print("Error executing the query", ex)
                
    def obtenerDato(self):
        row = self.tablaDeleteRecord.currentRow()
        nombre_cliente = self.tablaDeleteRecord.item(row, 1).text()
        self.nombre_buscar3.setText(nombre_cliente)
        # self.tablaDeleteRecord.clearSelection()  # Deseleccionar la fila eliminada
    
    def delete(self): # VER PORQUE NO DESHABILITA      
        if not self.tablaDeleteRecord.currentItem():
            mensaje_ingreso_datos("Registro de cliente","Debe seleccione el registro de la tabla y presione 'ELIMINAR'")
            return
        
        selectedRow = self.tablaDeleteRecord.currentItem().row()
        idUser = int(self.tablaDeleteRecord.item(selectedRow, 0).text())

        elim = inicio("Registro de cliente","¿Desea eliminar el cliente?")
        if elim == QMessageBox.StandardButton.Yes:
            elim2 = inicio("Registro de cliente","¿Estas seguro de eliminar cliente?")
            if elim2 == QMessageBox.StandardButton.Yes:
                try:
                    db = conectar_base_de_datos()
                    cursor = db.cursor()
                    cursor.execute("UPDATE usuario SET habilitado = 0 WHERE id_usuario = %s", (idUser,))
                    # cursor.execute(f"UPDATE disciplina SET habilitado = 0 WHERE id_disciplina= {id_dis}")#, (id_dis,))
                    # cursor.execute(f"DELETE FROM usuario WHERE id_usuario = {idUser} AND habilitado = 1")    
                    db.commit()
                    if cursor:
                        self.tablaDeleteRecord.removeRow(selectedRow)
                        self.nombre_buscar3.clear()
                        self.limpiar_tabla()
                        self.tablaDeleteRecord.clearSelection()
                        ingreso_datos("Registro de cliente","Registo eliminado")
                    else:
                        ingreso_datos("Registro de cliente","Registo no eliminado")
                        self.tablaDeleteRecord.clearSelection()  # Deseleccionar la fila eliminada
                        
                    cursor.close()
                    db.close()

                except Error as ex:
                    errorConsulta("Registro de cliente",f"Error en la consulta: {str(ex)}")
                    print("Error executing the query", ex)
            else:
                print("No se elimina registro")
        else:
            print("Decidio no elimina registro")

    def limpiar_tabla(self):
        borrarTabla(self) 
    
    # -------------- PESTAÑA DE DISCIPLINA -----------------
    def guardarACTIV(self):
        actividad = self.input_disciplina4.text().capitalize()
        precio = self.input_precio.text()
        
        validacion = guardarACTIVIDAD(self, actividad, precio)
        if validacion != "Carga de datos correcta":
            mensaje_ingreso_datos("Error de validación", "Verifique los datos por favor")
            return
        
        try:
            db = conectar_base_de_datos()
            cursor = db.cursor()

            # Verificar si la actividad ya existe
            cursor.execute("SELECT COUNT(*) FROM disciplina WHERE nombre = %s AND habilitado = 1", (actividad,))
            resultado = cursor.fetchone()

            if resultado[0] > 0:
                # Si la actividad ya existe, mostrar un mensaje y no insertar
                ingreso_datos("Registro de disciplina", "La actividad ya está registrada")
            else:
                # Si la actividad no existe, insertar la nueva actividad
                cursor.execute("INSERT INTO disciplina (nombre, precio) VALUES (%s, %s)", (actividad, precio))
                db.commit()

                if cursor.rowcount > 0:
                    ingreso_datos("Registro de disciplina", "Registro cargado")
                    self.input_disciplina4.clear()
                    self.input_precio.clear()
                else:
                    ingreso_datos("Registro de disciplina", "Registro no cargado")

            cursor.close()
            db.close()
        except Error as ex:
            errorConsulta("Registro de disciplina", f"Error en la consulta: {str(ex)}")
            print("Error executing the query", ex)

    def mostrarACTIC(self):
        try:
            db = conectar_base_de_datos()
            cursor = db.cursor()
            cursor.execute("SELECT id_disciplina, nombre, precio AS 'PRECIO ($)' FROM disciplina WHERE disciplina.habilitado = 1 ORDER BY id_disciplina")
            resultados = cursor.fetchall()

            if len(resultados) > 0:
                tabla_DISCIPLINA(self, resultados, cursor, QHeaderView, QTableWidget, QAbstractItemView, QTableWidgetItem, Qt)
                self.tableActivi.clearSelection()  # Deseleccionar la fila eliminada
            else:
                ingreso_datos("Registro de disciplina","Tabla sin registros")
                    
            cursor.close()  
            db.close()
        
        except Error as ex:
            errorConsulta("Registro de disciplina",f"Error en la consulta: {str(ex)}")
            print("Error executing the query", ex)
        
    def seleccionDeDatos(self):
        completar_CAMPOS_ACTIVIDAD(self)
        
    def actualizarACTIV(self):
        if not self.tableActivi.currentItem():
            mensaje_ingreso_datos("Registro de disciplina","Por favor seleccione un registro para actualizar")
            return
        
        id_dis = int(self.tableActivi.item(self.tableActivi.currentRow(), 0).text())
        actividad = self.input_disciplina4.text().capitalize().title()
        precio = self.input_precio.text().replace(".","")
        
        patronA = re.compile(r'^[a-zA-ZáéíóúÁÉÍÓÚüÜ\'\s]+$') 
        if not isinstance(actividad, str) or actividad.isspace() or not patronA.match(actividad):
            mensaje_ingreso_datos("Registro de disciplina","Debe elegir una disciplina")
            return

        patron2 = re.compile(r'^[0-9]+$')
        if not (precio.isdigit() and patron2.match(precio)):# and len(precio) > 0):
            mensaje_ingreso_datos("Registro de disciplina","El precio debe ser número entero. Sin coma ','.")
            return
        if precio: 
            precio = int(precio)
        
        try:
            db = conectar_base_de_datos()
            cursor = db.cursor()
            cursor.execute("UPDATE disciplina SET nombre=%s, precio=%s WHERE id_disciplina=%s", (actividad, precio, id_dis))
            db.commit()
            
            if cursor:
                ingreso_datos("Registro de disciplina","Registro actualizado")
                self.input_disciplina4.clear()
                self.input_precio.clear()
                self.mostrarACTIC()
                self.tableActivi.clearSelection()  # Deseleccionar la fila
            else:
                ingreso_datos("Registro de disciplina","Registro no actualizado")
                
            cursor.close()
            db.close()

        except Error as ex:
            errorConsulta("Registro de disciplina",f"Error en la consulta: {str(ex)}")
            print("Error al ejecutar la consulta", ex)
            
    def limpiar_tabla_disciplina(self):
        clear_tabla_disciplina(self)
    
    def limp(self):
        self.input_disciplina4.clear()
        self.input_precio.clear()
    
    def eliminarACTIV(self):#-----------------------------
        # Primero corroborar la seleccion de la fila
        if not self.tableActivi.currentItem():
            mensaje_ingreso_datos("Registro de disciplina","Debe seleccione el registro de la tabla y presione 'ELIMINAR'")
            return
        
        # Selecciona la fila acutal
        selectedRow = self.tableActivi.currentItem().row()
        id_dis = int(self.tableActivi.item(selectedRow, 0).text())
        
        # Verifica si ya se presionó el botón sin seleccionar una nueva fila
        if self.eliminacion_pendiente:
            mensaje_ingreso_datos("Registro de empleado", "Debe seleccionar una nueva fila antes de eliminar otro registro")
            return
        
        elim = inicio("Registro de disciplina","¿Desea eliminar la disciplina?")
        if elim == QMessageBox.StandardButton.Yes:
            elim2 = inicio("Registro de disciplina","¿Seguro que desea eliminar la disciplina?")
            if elim2 == QMessageBox.StandardButton.Yes:
                try:
                    db = conectar_base_de_datos()
                    cursor = db.cursor()
                    # cursor.execute(f"DELETE FROM disciplina WHERE id_disciplina = {id_dis}")    
                    cursor.execute("UPDATE disciplina SET habilitado = 0 WHERE id_disciplina= %s", (id_dis,))
                    if cursor:
                        ingreso_datos("Registro de disciplina","Registo eliminado")
                        self.tableActivi.removeRow(selectedRow)
                        self.input_disciplina4.clear()
                        self.input_precio.clear()
                        self.tableActivi.clearSelection()  # Deseleccionar la fila eliminada
                    else:
                        ingreso_datos("Registro de disciplina","Registo no eliminado")
                        
                    cursor.close()
                    db.commit()
                    self.eliminacion_pendiente = True  # Marca que se ha eliminado un registro y no se ha seleccionado una nueva fila

                except Error as ex:
                    errorConsulta("Registro de disciplina",f"Error en la consulta: {str(ex)}")
            else:
                print("NO se elimina disciplina")
        else:
            print("NO elimina disciplina")
            
    def planilla_disciplina(self):
        tabla_registroDISCIPLINA(self,Workbook,Font,PatternFill,Border,Side,numbers,QFileDialog)
       
    # ----------- PESTAÑA PAGOS -----------------
    # Limpiar campos
    def camposLimpios(self):
        self.idUser.clear()
        self.idDis.clear()
        self.input_tipoDePago.clear()
        self.input_fechaDePago.setDate(QDate.currentDate())
        
    def actualizar_precio(self, text):
        if text in self.disciplinas:
            id_disciplina, precio = self.disciplinas[text]
            self.label_monto.setText(f"Precio: $ {precio}")
            self.current_id_disciplina = id_disciplina
        else:
            self.label_monto.setText("Precio no encontrado")
            
    def guardarPagos(self):
        id_alumno = self.idUser.text()
        # Verificar si el id_alumno está en las sugerencias
        if not id_alumno in self.sugerencia:
            mensaje_ingreso_datos("Registro de pagos", "Debe elegir un DNI de usuario de la lista de sugerencias")
            return
        
        id_activ = self.current_id_disciplina
        if id_activ is None:
            mensaje_ingreso_datos("Registro de pagos", "Debe seleccionar una disciplina antes de guardar el pago")
            return
        
        tipo = self.input_tipoDePago.text()
        date = self.input_fechaDePago.date().toPyDate()
        monto = self.label_monto.text().split(": $ ")[1] 
        monto = int(monto)
        
        patronB = re.compile(r'^[a-zA-ZáéíóúÁÉÍÓÚüÜ\'\s]+$') 
        if not isinstance(tipo, str) or not patronB.match(tipo):
            mensaje_ingreso_datos("Registro de pagos","Debe elegir un tipo de pago")
            return
        
        medio_de_pago = ["Efectivo", "Transferencia"]
        if tipo not in medio_de_pago:
            mensaje_ingreso_datos("Registro de pagos","Debe elegir entre 'Efectivo' o 'Transferencia'.")
            return
        elif tipo in medio_de_pago:
            tipo = tipo.capitalize()
            
        try:
            db = conectar_base_de_datos()
            cursor = db.cursor()
            
                # Verificar si ya existe un registro con los mismos id_usuario, id_disciplina y fecha
            cursor.execute("SELECT COUNT(*) FROM pago WHERE id_usuario = %s AND id_disciplina = %s AND modalidad = %s AND fecha = %s", (id_alumno, id_activ, tipo, date))
            result = cursor.fetchone()
            if result[0] > 0:
                mensaje_ingreso_datos("Registro de pagos", "Ya existe un registro con los mismos datos")
                return
            
            cursor.execute("INSERT INTO pago (id_usuario, id_disciplina, modalidad, fecha, precio) VALUE (%s, %s, %s, %s, %s)", (id_alumno, id_activ, tipo, date, monto))
            db.commit()
            if cursor:
                ingreso_datos("Registro de pagos","Registro cargado")
                self.camposLimpios()
                self.label_monto.clear()
            else:
                ingreso_datos("Registro de pagos","Registro no cargado")
                
            cursor.close()
            db.close()
        except Error as ex:
            errorConsulta("Registro de pagos",f"Error en la cosulta: {str(ex)}")
            print(ex)

    def mostrarPagos(self):
        try:
            db = conectar_base_de_datos()
            cursor = db.cursor()
            cursor.execute("SELECT p.id_pago, p.id_usuario, c.nombre AS DISCIPLINA, p.modalidad, p.fecha, p.precio AS 'PRECIO ($)' FROM pago as p INNER JOIN disciplina as c on p.id_disciplina = c.id_disciplina WHERE p.habilitado = 1 ORDER BY p.fecha;")
            result = cursor.fetchall()
            
            if len(result) > 0:
                tabla_pagos(self, cursor, result, QHeaderView, QTableWidget, QAbstractItemView, QTableWidgetItem, QDate, Qt)
                               
                self.tablePagos.clearSelection()  # Deseleccionar la fila eliminada
            else:
                ingreso_datos("Registro de pagos","Tabla se encuentra vacia")   
            cursor.close()
            db.close()
            
        except Error as ex:
            errorConsulta("Registro de pagos",f"Error en la cosulta: {str(ex)}")
        
    def establecer_datos(self):
        seleccionDeTablaPAGOS(self,QDate)
        
    def actualizarPagos(self):
        if not self.tablePagos.currentItem():
            mensaje_ingreso_datos("Registro de pago","Por favor seleccione un registro para actualizar")
            return
        
        idpago = int(self.tablePagos.item(self.tablePagos.currentRow(),0).text())
        id_alumno = self.idUser.text() #currentData()[0]
        # Verificar si el id_alumno está en las sugerencias
        if id_alumno not in self.sugerencia:
            mensaje_ingreso_datos("Registro de pagos", "Debe elegir un ID de usuario de la lista de sugerencias")
            return
        
        id_activ = self.current_id_disciplina
        if id_activ is None:
            mensaje_ingreso_datos("Registro de pagos", "Debe seleccionar una disciplina antes de guardar el pago")
            return
        
        tipo = self.input_tipoDePago.text()
        date = self.input_fechaDePago.date().toPyDate()
        monto = self.label_monto.text().split(": $ ")[1] 
            
        patronB = re.compile(r'^[a-zA-ZáéíóúÁÉÍÓÚüÜ\'\s]+$') 
        if not isinstance(tipo, str) or not patronB.match(tipo):
            mensaje_ingreso_datos("Registro de pagos","Debe elegir un tipo de pago")
            return
        
        medio_de_pago = ["Efectivo", "Transferencia"]
        if tipo not in medio_de_pago:
            mensaje_ingreso_datos("Registro de pagos","Debe elegir un sexo entre 'Efectivo' o 'Transferencia'.")
            return
        elif tipo in medio_de_pago:
            tipo = tipo.capitalize()
            
        try:
            db = conectar_base_de_datos()
            cursor = db.cursor()
            cursor.execute(f"UPDATE pago SET id_usuario='{id_alumno}', id_disciplina='{id_activ}', modalidad='{tipo}', fecha='{date}', precio='{monto}' WHERE id_pago='{idpago}'")#
            db.commit()
        
            if cursor:
                ingreso_datos("Registro de pagos","Registro actualizado")
                self.camposLimpios()
                self.mostrarPagos()
                self.label_monto.clear()
            else:
                ingreso_datos("Registro de pagos","Registro no actualizado")
            cursor.close()
            db.close()
            self.tablePagos.clearSelection()  # Deseleccionar la fila eliminada 
        except Error as ex:
            errorConsulta("Registro de pagos",f"Error en la cosulta: {str(ex)}")
    
    def eliminarPagos(self):#--------------------- 
        # Primero corroborar la seleccion de la fila
        if not self.tablePagos.currentItem():
            mensaje_ingreso_datos("Registro de pagos","Debe seleccione el registro de la tabla y presione 'ELIMINAR'")
            return
        
        # Selecciona la fila acutal
        selectedRow = self.tablePagos.currentItem().row()
        itemPagos = int(self.tablePagos.item(selectedRow, 0).text())
        
        # Verifica si ya se presionó el botón sin seleccionar una nueva fila
        if self.eliminacion_pendiente:
            mensaje_ingreso_datos("Registro de pagos", "Debe seleccionar una nueva fila antes de eliminar otro registro")
            return
        
        elimPagos = inicio("Registro de pagos","¿Desea eliminar el pago?")
        if elimPagos == QMessageBox.StandardButton.Yes:
            elimPagos2 = inicio("Registro de pagos","¿Seguro que desea eliminar el pago?")
            if elimPagos2 == QMessageBox.StandardButton.Yes:
                try:
                    db = conectar_base_de_datos()
                    cursor = db.cursor()
                    cursor.execute("UPDATE pago SET habilitado = 0 WHERE id_pago = %s", (itemPagos,))
                    # cursor.execute(f"DELETE FROM pago WHERE id_pago = '{itemPagos}'")    
                    db.commit()
                    
                    if cursor:
                        ingreso_datos("Registro de pagos","Registo eliminado")
                        self.tablePagos.removeRow(selectedRow)
                        self.camposLimpios()
                    
                        self.tablePagos.clearSelection()  # Deseleccionar la fila eliminada 
                    else:
                        ingreso_datos("Registro de pagos","Registo no eliminado")
                                        
                    cursor.close()
                    db.close()
                    
                except Error as ex:
                    errorConsulta("Registro de pagos",f"Error en la consulta: {str(ex)}")
                cursor.close()
                db.close()
                self.eliminacion_pendiente = True  # Marca que se ha eliminado un registro y no se ha seleccionado una nueva fila
                
            else:
                print("NO se elimina pago")
        else:
            print("NO elimina pago")
            
    def planilla_pagos(self):
        pagos_EXCEL(self,Workbook,Font,PatternFill,Border,Side,numbers,QFileDialog)
            
    # ----------- PESTAÑA BALANCE -----------------
    def consultar_TotalPorAlumno(self):
        dni = self.view_nomb.text()

        patron = re.compile(r'^[0-9]+$')
        if not isinstance(dni, str)  or not patron.match(dni):
            mensaje_ingreso_datos("Balances","Debe elige un DNI")
            return
        if dni not in self.numeros_dni:  # Verificar si el DNI ingresado está en la lista de DNIs
            mensaje_ingreso_datos("Balances", "El DNI ingresado no está registrado.")
            return
        
        fecha_inicio = self.view_fechaDePago.date().toString("yyyy-MM-dd")
        
        if not fecha_inicio or fecha_inicio > QDate.currentDate().toString("yyyy-MM-dd"):
            self.view_fechaDePago.setDate(QDate.currentDate())
            mensaje_ingreso_datos("Balances","La fecha de inicio debe ser la actual o una fecha anterior.")
            return 
        
        fecha_fin = self.view_al2.date().toString("yyyy-MM-dd")

        if fecha_fin > QDate.currentDate().toString("yyyy-MM-dd") or fecha_fin < fecha_inicio:
            self.view_al2.setDate(QDate.currentDate())
            mensaje_ingreso_datos("Balances","La fecha de inicio debe ser la actual o una fecha anterior.\nLa fecha final de periodo debe ser igual o no anterior a la fecha de inicio de periodo")
            return
        
        try:
            db = conectar_base_de_datos()
            cursor = db.cursor()
            query = f"SELECT u.nombre, u.apellido, u.dni, u.sexo, u.edad, u.fecha_registro AS REGISTRO, d.nombre AS DISCIPLINA, p.precio AS 'PRECIO ($)', p.fecha, p.modalidad FROM usuario u JOIN pago p ON u.dni = p.id_usuario JOIN disciplina d ON p.id_disciplina = d.id_disciplina WHERE u.dni = '{dni}' AND p.fecha BETWEEN '{fecha_inicio}' AND '{fecha_fin}' ORDER BY p.fecha ASC"
            cursor.execute(query)
            results = cursor.fetchall()
            
            if len(results) > 0:
                ingreso_datos("Balances",f"Se encontraron {len(results)} coincidencias.")
                self.view_fechaDePago.setDate(QDate.currentDate())
                self.view_al2.setDate(QDate.currentDate())
                self.view_nomb.clear()
                consultaPorAlumno(self,cursor,results,QHeaderView,QTableWidget,QAbstractItemView,QTableWidgetItem,QDate,Qt)
                            
            else:
                ingreso_datos("Balances",f"Se encontraron {len(results)} coincidencias.")
                self.view_fechaDePago.setDate(QDate.currentDate())
                self.view_al2.setDate(QDate.currentDate())
                self.view_nomb.clear()
                
            cursor.close()
            db.close()
            
        except Error as ex:
            errorConsulta("Balances",f"Error en la consulta: {str(ex)}")
            print("Error executing the query", ex)
    
    def consultar_TotalAlumno(self):
        fecha_inicio = self.view_fechaDePago.date().toString("yyyy-MM-dd")
        if not fecha_inicio or fecha_inicio > QDate.currentDate().toString("yyyy-MM-dd"):
            self.view_fechaDePago.setDate(QDate.currentDate())
            mensaje_ingreso_datos("Balances","La fecha de inicio debe ser la actual o una fecha anterior.")
            return 
        
        fecha_fin = self.view_al2.date().toString("yyyy-MM-dd")
        if fecha_fin > QDate.currentDate().toString("yyyy-MM-dd") or fecha_fin < fecha_inicio:
            self.view_al2.setDate(QDate.currentDate())
            mensaje_ingreso_datos("Balances","La fecha de inicio debe ser la actual o una fecha anterior.\nLa fecha final de periodo debe ser igual o no anterior a la fecha de inicio de periodo")
            return
        
        try:
            db = conectar_base_de_datos()
            cursor = db.cursor()
            query = f"SELECT u.nombre, u.apellido, u.dni, u.sexo, u.edad, u.celular, u.fecha_registro AS REGISTRO, d.nombre AS DISCIPLINA, p.precio AS 'PRECIO ($)', p.fecha, p.modalidad FROM usuario u JOIN pago p ON u.dni = p.id_usuario JOIN disciplina d ON p.id_disciplina = d.id_disciplina WHERE p.fecha BETWEEN '{fecha_inicio}' AND '{fecha_fin}' ORDER BY p.fecha ASC"
            cursor.execute(query)
            results = cursor.fetchall()
            
            if results:
                ingreso_datos("Balances",f"Se encontraron {len(results)} coincidencias.")
                self.view_fechaDePago.setDate(QDate.currentDate())
                self.view_al.setDate(QDate.currentDate())
                totalAlumno(self,cursor,results,QHeaderView,QTableWidget,QAbstractItemView,QAbstractScrollArea,QTableWidgetItem,QDate,Qt)
                
            else: 
                ingreso_datos("Balances",f"Se encontraron {len(results)} coincidencias.")
                self.view_fechaDePago.setDate(QDate.currentDate())
                self.view_al2.setDate(QDate.currentDate())
                
            cursor.close()
            db.close()
                
        except Error as ex:
            errorConsulta("Balances",f"Error en la consulta: {str(ex)}")
            print("Error executing the query", ex)
            
    def limpiar_tabla_balance(self):
        limpiar(self)
            
    def consultar_TotalDisciplina(self):
        fecha_inicio = self.view_fechaDePago.date().toString("yyyy-MM-dd")
        if not fecha_inicio or fecha_inicio > QDate.currentDate().toString("yyyy-MM-dd"):
            self.view_fechaDePago.setDate(QDate.currentDate())
            mensaje_ingreso_datos("Balances","La fecha de inicio debe ser la actual o una fecha anterior.")
            return 
        
        fecha_fin = self.view_al2.date().toString("yyyy-MM-dd")
        if fecha_fin > QDate.currentDate().toString("yyyy-MM-dd") or fecha_fin < fecha_inicio:
            self.view_al2.setDate(QDate.currentDate())
            mensaje_ingreso_datos("Balances","La fecha de inicio debe ser la actual o una fecha anterior.\nLa fecha final de periodo debe ser igual o no anterior a la fecha de inicio de periodo")
            return
        try:
            db = conectar_base_de_datos()
            cursor = db.cursor()
            query = f"SELECT u.nombre, u.apellido, d.nombre AS DISCIPLINA, SUM(p.precio) AS 'TOTAL PRECIO ($)', '{fecha_inicio}' AS inicio_periodo, '{fecha_fin}' AS fin_periodo FROM usuario u JOIN pago p ON u.dni = p.id_usuario JOIN disciplina d ON p.id_disciplina = d.id_disciplina WHERE p.fecha BETWEEN '{fecha_inicio}' AND '{fecha_fin}' GROUP BY u.nombre, u.apellido, d.nombre"
            cursor.execute(query)
            results = cursor.fetchall()
            
            if results:
                ingreso_datos("Balances",f"Se encontraron {len(results)} coincidencias.")
                self.view_fechaDePago.setDate(QDate.currentDate())
                self.view_al2.setDate(QDate.currentDate())
                consultarDeDisciplina(self,cursor,results,QHeaderView,QTableWidget,QAbstractItemView,QTableWidgetItem,QDate,Qt)
                            
            else:
                ingreso_datos("Balances",f"Se encontraron {len(results)} coincidencias.")
                self.view_fechaDePago.setDate(QDate.currentDate())
                self.view_al2.setDate(QDate.currentDate())
                
            cursor.close()
            db.close()
        except Error as ex:
            errorConsulta("Balances",f"Error en la consulta: {str(ex)}")
            print("Error executing the query", ex)
            
    def consultar_PorDisciplina(self):
        actividad = self.view_disciplina.text()
        
        patrones = re.compile(r'^[a-zA-ZáéíóúÁÉÍÓÚüÜ\'\s]+$')
        if not isinstance(actividad,str) or not patrones.match(actividad):
            mensaje_ingreso_datos("Balances","Debe elegir una disciplina")
            return
        if actividad not in self.activi:  # Verificar si el DNI ingresado está en la lista de DNIs
            mensaje_ingreso_datos("Balances", "La disciplina ingresada no está registrado.")
            self.view_disciplina.clear()
            return
        
        fecha_inicio = self.view_fechaDePago.date().toString("yyyy-MM-dd")
        if not fecha_inicio or fecha_inicio > QDate.currentDate().toString("yyyy-MM-dd"):
            self.view_fechaDePago.setDate(QDate.currentDate())
            mensaje_ingreso_datos("Balances","La fecha de inicio debe ser la actual o una fecha anterior.")
            return 
        
        fecha_fin = self.view_al2.date().toString("yyyy-MM-dd")
        if fecha_fin > QDate.currentDate().toString("yyyy-MM-dd") or fecha_fin < fecha_inicio:
            self.view_al2.setDate(QDate.currentDate())
            mensaje_ingreso_datos("Balances","La fecha de inicio debe ser la actual o una fecha anterior.\nLa fecha final de periodo debe ser igual o no anterior a la fecha de inicio de periodo")
            return
                
        try:
            db = conectar_base_de_datos()
            cursor = db.cursor()
            query = f"SELECT u.nombre, u.apellido, u.dni, u.sexo, u.edad, d.nombre AS DISCIPLINA, p.modalidad, p.fecha, SUM(p.precio) AS 'TOTAL PRECIO ($)' FROM usuario u JOIN pago p ON u.dni = p.id_usuario JOIN disciplina d ON p.id_disciplina = d.id_disciplina WHERE p.fecha BETWEEN '{fecha_inicio}' AND '{fecha_fin}' AND d.nombre = '{actividad}' GROUP BY u.nombre, u.apellido, u.dni, u.sexo, u.edad, d.nombre, p.modalidad, p.fecha ORDER BY p.fecha ASC"
            
            cursor.execute(query)
            results = cursor.fetchall()
                    
            if len(results) > 0:
                ingreso_datos("Balances",f"Se encontraron {len(results)} coincidencias.")
                self.view_disciplina.clear()
                self.view_fechaDePago.setDate(QDate.currentDate())
                self.view_al2.setDate(QDate.currentDate())
                consultaPorDisciplina(self,cursor,results,QHeaderView,QTableWidget,QAbstractItemView,QTableWidgetItem,QDate,Qt)
            else:
                ingreso_datos("Balances",f"Se encontraron {len(results)} coincidencias.")
                self.view_fechaDePago.setDate(QDate.currentDate())
                self.view_al2.setDate(QDate.currentDate())
                self.view_disciplina.clear()
                
            cursor.close()
            db.close()
                
        except Error as ex:
            errorConsulta("Balances",f"Error en la consulta: {str(ex)}")
            print("Error executing the query", ex)
            
    def consultar_AsisteTotal(self):
        fecha_inicio2 = self.view_asistencia.date().toString("yyyy-MM-dd")
        if not fecha_inicio2 or fecha_inicio2 > QDate.currentDate().toString("yyyy-MM-dd"):
            self.view_asistencia.setDate(QDate.currentDate())
            mensaje_ingreso_datos("Balances","La fecha de inicio debe ser la actual o una fecha anterior.")
            return 
        
        fecha_fin2 = self.view_al.date().toString("yyyy-MM-dd")
        if fecha_fin2 > QDate.currentDate().toString("yyyy-MM-dd") or fecha_fin2 < fecha_inicio2:
            self.view_al.setDate(QDate.currentDate())
            mensaje_ingreso_datos("Balances","La fecha de inicio debe ser la actual o una fecha anterior.\nLa fecha final de periodo debe ser igual o no anterior a la fecha de inicio de periodo")
            return     
        
        try:
            db = conectar_base_de_datos()
            cursor = db.cursor()            
            cursor.execute(f"SELECT u.nombre, u.apellido, u.dni, u.sexo, u.edad, a.asistencia FROM usuario u JOIN asistencia a ON u.dni = a.dni WHERE a.asistencia BETWEEN '{fecha_inicio2}' AND '{fecha_fin2}' ORDER BY a.asistencia ASC")
            results = cursor.fetchall()
            
            if len(results) > 0:
                ingreso_datos("Balances",f"Se encontraron {len(results)} coincidencias.")
                self.view_asistencia.setDate(QDate.currentDate())
                self.view_al.setDate(QDate.currentDate())
                asistenciaTotal(self,cursor,results,QHeaderView,QTableWidget,QAbstractItemView,QTableWidgetItem,QDate,Qt)
                
            else:
                ingreso_datos("Balances",f"Se encontraron {len(results)} coincidencias.")
                self.view_asistencia.setDate(QDate.currentDate())
                self.view_al.setDate(QDate.currentDate())
                
            cursor.close()
            db.close()
                
        except Error as ex:
            errorConsulta("Balances",f"Error en la consulta: {str(ex)}")
            print("Error executing the query", ex)
    
    def consultar_AsistePorAlumno(self):
        alumno = self.view_nomb.text()
        
        patron = re.compile(r'^[0-9]+$')
        if not isinstance(alumno, str) or not patron.match(alumno): 
            mensaje_ingreso_datos("Balances","Debe elige un DNI")
            return
        if alumno not in self.numeros_dni:  # Verificar si el DNI ingresado está en la lista de DNIs
            mensaje_ingreso_datos("Balances", "El DNI ingresado no está registrado.")
            self.view_nomb.clear()
            return
        
        fecha_inicio = self.view_asistencia.date().toString("yyyy-MM-dd")
        if not fecha_inicio or fecha_inicio > QDate.currentDate().toString("yyyy-MM-dd"):
            self.view_asistencia.setDate(QDate.currentDate())
            mensaje_ingreso_datos("Balances","La fecha de inicio debe ser la actual o una fecha anterior.")
            return
        
        fecha_fin = self.view_al.date().toString("yyyy-MM-dd")
        if fecha_fin > QDate.currentDate().toString("yyyy-MM-dd") or fecha_fin < fecha_inicio:
            self.view_al.setDate(QDate.currentDate())
            mensaje_ingreso_datos("Balances","La fecha de inicio debe ser la actual o una fecha anterior.\nLa fecha final de periodo debe ser igual o no anterior a la fecha de inicio de periodo")
            return
        
        try:
            db = conectar_base_de_datos()
            cursor = db.cursor()
            
            # Ejecutar la consulta
            cursor.execute(f"SELECT u.nombre, u.apellido, u.dni, u.sexo, u.edad, a.asistencia FROM usuario u JOIN asistencia a ON u.dni = a.dni WHERE a.asistencia BETWEEN '{fecha_inicio}' AND '{fecha_fin}' AND u.dni = '{alumno}' ORDER BY a.asistencia ASC")
            results5 = cursor.fetchall()

            if len(results5) > 0:
                ingreso_datos("Balances",f"Se encontraron {len(results5)} asistencias.")
                self.view_nomb.clear()
                self.view_asistencia.setDate(QDate.currentDate())
                self.view_al.setDate(QDate.currentDate())
                asistenciaPorAlumno(self,cursor,results5,QHeaderView,QTableWidget,QAbstractItemView,QTableWidgetItem,QDate,Qt)
            else:
                ingreso_datos("Balances",f"Se encontraron {len(results5)} asistencias.")
                self.view_asistencia.setDate(QDate.currentDate())
                self.view_al.setDate(QDate.currentDate())
                self.view_nomb.clear()
                
            cursor.close()
            db.close()
                
        except Error as ex:
            errorConsulta("Balances",f"Error en la consulta: {str(ex)}")
            print("Error executing the query", ex) 
        
    def tabla_balance(self):
        excelConsulta(self,Workbook,Font,PatternFill,Border,Side,numbers,QFileDialog)
            
    # ----------------------- EMPLEADOS Y HORAS -------------------------------------
    def guardar_id_empleado(self, texto):
        # Buscar el id_empleado correspondiente al nombre seleccionado
        for nombre, id_empleado in self.listas:
            if nombre == texto:
                # Guardar el id_empleado en una variable o tabla, según tu necesidad
                self.id_empleado_seleccionado = id_empleado
                break
        else:
            self.id_empleado_seleccionado = None
            
    def guardar_horas(self):
        id_empleado_seleccionado = self.id_empleado_seleccionado  # Obtener el id_empleado seleccionado
        # Verificar si el id_alumno está en las sugerencias
        if id_empleado_seleccionado is None:
            mensaje_ingreso_datos("Registro de horas", "Debe elegir un Nombre de empleado de la lista de sugerencias")
            return
        
        horas_horas = self.horas_tra.text()
        fecha_horas = self.fecha_tra.date().toPyDate()
    
        patron_mun = re.compile(r'^[0-9]+$')
        if not (horas_horas.isnumeric() and patron_mun.match(horas_horas)):
            mensaje_ingreso_datos("Registro de hora","Las 'Horas diarias' debe ser numérico.")
            return
        horas_horas = int(horas_horas)
        
        try:
            db = conectar_base_de_datos()
            cursor = db.cursor()
            
            # Verificar si ya existe un registro con los mismos id_usuario, id_disciplina y fecha
            cursor.execute("SELECT COUNT(*) FROM hora WHERE id_empleado = %s AND horas_diaria = %s AND fecha = %s", (id_empleado_seleccionado, horas_horas, fecha_horas))
            result = cursor.fetchone()
            if result[0] > 0:
                ingreso_datos("Registro de horas", "Ya existe un registro con los mismos datos")
                return
            
            cursor.execute("INSERT INTO hora (id_empleado,horas_diaria,fecha) VALUES (%s,%s,%s)", (id_empleado_seleccionado,horas_horas,fecha_horas))
            db.commit()
            
            if cursor:
                ingreso_datos("Registro de horas","Registro cargado")
                self.id_horas_empleado.clear()
                self.horas_tra.clear()
                self.fecha_tra.setDate(QDate.currentDate())
            else:
                ingreso_datos("Registro de horas","Registro no cargado")
                
            cursor.close()
            db.close()
        except Error as ex:
            errorConsulta("Registro de horas",f"Error en la consulta: {str(ex)}")
            print("Error executing the query", ex)
                
    def limpiar_tabla_horas(self):
        clearTabla(self) 
        
    def horas_empleado(self):        
        id_empleado = self.id_horas_empleado.text()
        if not id_empleado:
            mensaje_ingreso_datos("Registro de horas","Debe completar el nombre del empleado")
            return
        
        principio = self.periodo.date().toString("yyyy-MM-dd")
        if not principio or principio > QDate.currentDate().toString("yyyy-MM-dd"):
            self.periodo.setDate(QDate.currentDate())
            mensaje_ingreso_datos("Registro de horas","La fecha de inicio debe ser la actual o una fecha anterior.")
            return 
        
        final = self.fin_tra.date().toString("yyyy-MM-dd")
        if final > QDate.currentDate().toString("yyyy-MM-dd") or final < principio:
            self.fin_tra.setDate(QDate.currentDate())
            mensaje_ingreso_datos("Registro de horas","La fecha de inicio debe ser la actual o una fecha anterior.\nLa fecha final de periodo debe ser igual o no anterior a la fecha de inicio de periodo")
            return
        
        try:
            db = conectar_base_de_datos()
            cursor = db.cursor()
            cursor.execute(f"SELECT h.id_hora AS REGISTRO, h.id_empleado, e.nombre, e.apellido, h.horas_diaria, h.fecha FROM hora AS h INNER JOIN registro_empleado AS e ON h.id_empleado = e.id_empleado WHERE h.id_empleado = '{id_empleado}' AND h.fecha BETWEEN '{principio}' AND '{final}' ORDER BY e.nombre, h.fecha")
            busqueda = cursor.fetchall()
                        
            if len(busqueda) > 0:
                ingreso_datos("Registro de horas",f"Se encontraron {len(busqueda)} coincidencias.")
                self.periodo.setDate(QDate.currentDate())
                self.fin_tra.setDate(QDate.currentDate())
                tabla_HorasXEmpleado(self,cursor,busqueda,QHeaderView,QTableWidget,QAbstractItemView,QTableWidgetItem,Qt,QDate)            
            else:    
                ingreso_datos("Registro de horas",f"Se encontraron {len(busqueda)} coincidencias.")
                self.periodo.setDate(QDate.currentDate())
                self.fin_tra.setDate(QDate.currentDate())
                
            cursor.close()
            db.close()
        except Error as ex:
            errorConsulta("Registro de horas",f"Error en la consulta: {str(ex)}")
            print("Error executing the query", ex)
        
    def horas_empleado_totales(self):
        principio = self.periodo.date().toString("yyyy-MM-dd")
        if not principio or principio > QDate.currentDate().toString("yyyy-MM-dd"):
            self.periodo.setDate(QDate.currentDate())
            mensaje_ingreso_datos("Registro de horas","La fecha de inicio debe ser la actual o una fecha anterior.")
            return 
        
        final = self.fin_tra.date().toString("yyyy-MM-dd")
        if final > QDate.currentDate().toString("yyyy-MM-dd") or final < principio:
            self.fin_tra.setDate(QDate.currentDate())
            mensaje_ingreso_datos("Registro de horas","La fecha de inicio debe ser la actual o una fecha anterior.\nLa fecha final de periodo debe ser igual o no anterior a la fecha de inicio de periodo")
            return
        
        try:
            db = conectar_base_de_datos()
            cursor = db.cursor()
            cursor.execute(f"SELECT h.id_hora AS REGISTRO, h.id_empleado, e.nombre, e.apellido, h.horas_diaria, h.fecha FROM hora AS h INNER JOIN registro_empleado AS e ON h.id_empleado = e.id_empleado AND h.fecha BETWEEN '{principio}' AND '{final}' ORDER BY h.id_empleado, h.fecha")
            busqueda = cursor.fetchall()
                        
            if len(busqueda) > 0:
                ingreso_datos("Registro de horas",f"Se encontraron {len(busqueda)} coincidencias.")
                self.periodo.setDate(QDate.currentDate())
                self.fin_tra.setDate(QDate.currentDate())
                tabla_HorasTotales(self,cursor,busqueda,QHeaderView,QTableWidget,QAbstractItemView,QTableWidgetItem,Qt,QDate)
            else:
                ingreso_datos("Registro de horas",f"Se encontraron {len(busqueda)} coincidencias.")
                self.periodo.setDate(QDate.currentDate())
                self.fin_tra.setDate(QDate.currentDate())
                
            cursor.close()
            db.close()
        except Error as ex:
            errorConsulta("Registro de horas",f"Error en la consulta: {str(ex)}")
            
    def autocompleto_de_datos_horas(self):
        autoCompletado(self,QDate,mensaje_ingreso_datos)
            
    def eliminar_horas(self): # ---------------
        # Primero corroborar la seleccion de la fila
        if not self.tablaHoras.currentItem():
            mensaje_ingreso_datos("Registro de horas","Debe buscar el registro a eliminar")
            return
        
        # Selecciona la fila acutal
        selectedRow = self.tablaHoras.currentRow()
        id_hor = int(self.tablaHoras.item(selectedRow, 0).text())
        
        # Verifica si ya se presionó el botón sin seleccionar una nueva fila
        if self.eliminacion_pendiente:
            mensaje_ingreso_datos("Registro de horas", "Debe seleccionar una nueva fila antes de eliminar otro registro")
            return
        
        elimHora = inicio("Registro de horas","¿Desea eliminar el registro?")
        if elimHora == QMessageBox.StandardButton.Yes:
            elimHora2 = inicio("Registro de horas","¿Seguro que desea eliminar el registro?")
            if elimHora2 == QMessageBox.StandardButton.Yes:
                try:
                    db = conectar_base_de_datos()
                    cursor = db.cursor()
                    cursor.execute("UPDATE hora SET habilitado = 0 WHERE id_hora = %s", (id_hor,))
                    # cursor.execute("DELETE FROM hora WHERE id_hora = %s", (id_hor,))
                    db.commit()
                    
                    if cursor:
                        ingreso_datos("Registro de horas","Registro eliminado")
                        self.tablaHoras.removeRow(selectedRow)

                        if self.tablaHoras.rowCount() == 1:
                            self.tablaHoras.setRowCount(0)  # Eliminar el registro de las sumatorias
                        
                        self.id_horas_empleado.clear()
                        self.horas_tra.clear()
                        self.fecha_tra.setDate(QDate.currentDate())
                        self.periodo.setDate(QDate.currentDate())
                        self.fin_tra.setDate(QDate.currentDate())
                        self.tablaHoras.clearSelection()  # Deseleccionar la fila eliminada
                    else:
                        ingreso_datos("Registro de horas","Registro no eliminado")
                        self.periodo.setDate(QDate.currentDate())
                        self.fin_tra.setDate(QDate.currentDate())
                        
                    cursor.close()
                    db.close()
                    self.eliminacion_pendiente = True  # Marca que se ha eliminado un registro y no se ha seleccionado una nueva fila
                    
                except Error as ex:
                    errorConsulta("Registro de horas",f"Error en la consulta: {str(ex)}")
                    print("Error executing the query", ex)
                cursor.close()
                db.close()
            else:
                print("NO se elimino")
        else:
            print("NO se eliimina hora")
                
    def actualizar_horas(self):
        # Verificar si se ha seleccionado una fila
        if not self.tablaHoras.currentItem():
            mensaje_ingreso_datos("Registro de horas","Debe seleccionar el registro de la tabla para actualizar")
            return
        
        id_ref = self.tablaHoras.item(self.tablaHoras.currentRow(), 0).text()
        id_ref = int(id_ref)
        id_empleado_seleccionado = self.id_empleado_seleccionado
        id_empleado_seleccionado = int(id_empleado_seleccionado) # Obtener el id_empleado seleccionado
        horas_h = self.horas_tra.text()
        fecha_h = self.fecha_tra.date().toPyDate()
            
        # Verificar si el id_alumno está en las sugerencias
        if id_empleado_seleccionado is None:
            mensaje_ingreso_datos("Registro de horas", "Debe elegir un Nombre de empleado de la lista de sugerencias")
            return
        
        patron_mun = re.compile(r'^[0-9]+$')
        if not (horas_h.isnumeric() and patron_mun.match(horas_h)):
            mensaje_ingreso_datos("Registro de horas","Las 'Horas diarias' debe ser numérico.")
            return
        horas_h = int(horas_h)
          
        try:
            db = conectar_base_de_datos()
            cursor = db.cursor()
            cursor.execute("UPDATE hora SET horas_diaria = %s, fecha = %s WHERE id_empleado = %s AND id_hora = %s", (horas_h, fecha_h, id_empleado_seleccionado, id_ref))
            db.commit() 
            
            if cursor:
                ingreso_datos("Registro de horas","Registro actualizado")
                self.id_horas_empleado.clear()
                self.horas_tra.clear()
                self.fecha_tra.setDate(QDate.currentDate())
            else:
                ingreso_datos("Registro de horas","Registro no actualizado")
                
            cursor.close()
            db.close() 
            
            self.tablaHoras.clearSelection() # Deselecciona la fila
            
        except Error as ex:
            errorConsulta("Registro de horas",f"Error en la consulta: {str(ex)}")
            print("Error executing the query", ex)
    
    def excel_horas(self):
        horas_Excel(self,Workbook,Font,PatternFill,Border,Side,numbers,QFileDialog)
        
    # -------------- LIBRO DIARIO -----------------------------
     
    def debe(self):    
        autocompletoDEBE(self)
    def haber(self):
        autocompletoHABER(self)
        
    def registrar_datos(self):        
        date = self.fecha_gastos.date().toPyDate()
        descripcion = self.concepto_debe.text().capitalize()
        descripcion_h = self.concepto_haber.text().capitalize()
        deber = self.debe.text()
        haberes = self.haber.text()
        
        # Lista de valores válidos para descripción y descripción_h
        listaDede = self.sugerencia1
        # print(listaDede)
        if not (descripcion in listaDede or descripcion == ""): 
            mensaje_ingreso_datos("Registro de Ingresos-Egresos", "El 'Debe', tiene que ser dentro de los sugerido o puede estar vacio.")
            return
        if descripcion in listaDede:
            descripcion_h.replace(descripcion_h,"")  #self.concepto_haber.setEnabled(False) and #self.concepto_haber.setText("")
        deber.replace(deber,"0")
        
        listaHaber = self.sugerencia2   
        # print(listaHaber) 
        if not (descripcion_h in listaHaber or descripcion_h == ""):
            mensaje_ingreso_datos("Registro de Ingresos-Egresos", "El 'Haber', tiene que ser dentro de los sugerido o puede estar vacio.")
            return
        if descripcion_h in listaHaber:
            descripcion.replace(descripcion,"") and self.debe.setText("0")#self.concepto_debe.setEnabled(False) and #self.concepto_debe.setText("")
        
        validadcion = validaciones(date,deber,haberes)#self,date,descripcion,descripcion_h,
        if validadcion != "Validación exitosa.":
            mensaje_ingreso_datos("Error de validación", "Verifique los datos por favor")
            return
        
        ingYegreso = inicio("Registro de Ingresos-Egresos","¿Desea guardar los datos?")
        if ingYegreso == QMessageBox.StandardButton.Yes: 
            try:
                db = conectar_base_de_datos()
                cursor = db.cursor()
                
                    # Verificar si ya existe un registro con los mismos id_usuario, id_disciplina y fecha
                cursor.execute("SELECT COUNT(*) FROM contabilidad WHERE fecha = %s AND concepto_debe = %s AND concepto_haber = %s AND debe = %s AND haber = %s", (date, descripcion,descripcion_h, deber, haberes))
                result = cursor.fetchone()
                if result[0] > 0:
                    mensaje_ingreso_datos("Registro de Ingresos-Egresos", "Ya existe un registro con los mismos datos")
                    return
                
                cursor.execute("INSERT INTO contabilidad (fecha, concepto_debe, concepto_haber, debe, haber) VALUES (%s, %s, %s, %s, %s)", 
                               (date, descripcion,descripcion_h, deber, haberes))
                db.commit()
                print(f"\nREGISTRO: {date}, {descripcion},{descripcion_h}, {deber}, {haberes}")
                if cursor:
                    ingreso_datos("Registro de Ingresos-Egresos","Datos cargados correctamente")
                    limpiarCampos(self,QDate)
                    self.concepto_debe.setEnabled(True)
                    self.concepto_haber.setEnabled(True)
                else:
                    ingreso_datos("Registro de Ingresos-Egresos","Datos no cargados correctamente")
                    
                cursor.close()
                db.close()
            except Error as ex:
                errorConsulta("Registro de Ingresos-Egresos",f"Error en la consulta: {str(ex)}")
                print("Error executing the query", ex)
        else:
            print("No se guarda registro")
            
    def selecionarTabla(self):
        selccionarTabla(self,mensaje_ingreso_datos,QDate)
        
    def actualizar_datos(self):
        if not self.tablaGastos.currentItem():
            mensaje_ingreso_datos("Registro de Ingresos-Egresos","Debe seleccionar el registro te la tabla para actualizar")
            return
        
        id_concepto = int(self.tablaGastos.item(self.tablaGastos.currentRow(),0).text())
        date = self.fecha_gastos.date().toPyDate()
        descripcion = self.concepto_debe.text().capitalize()
        descripcion_h = self.concepto_haber.text().capitalize()
        deber = self.debe.text()
        haberes = self.haber.text()
        
        # Lista de valores válidos para descripción y descripción_h
        listaDede = self.sugerencia1
        # print(listaDede)
        if not (descripcion in listaDede or descripcion == ""): 
            mensaje_ingreso_datos("Registro de Ingresos-Egresos", "El 'Debe', tiene que ser dentro de los sugerido o puede estar vacio.")
            return
        if descripcion in listaDede:
            descripcion_h.replace(descripcion_h,"")  #self.concepto_haber.setEnabled(False) and #self.concepto_haber.setText("")
        deber.replace(deber,"0")
        
        listaHaber = self.sugerencia2   
        # print(listaHaber) 
        if not (descripcion_h in listaHaber or descripcion_h == ""):
            mensaje_ingreso_datos("Registro de Ingresos-Egresos", "El 'Haber', tiene que ser dentro de los sugerido o puede estar vacio.")
            return
        if descripcion_h in listaHaber:
            descripcion.replace(descripcion,"") and self.debe.setText("0")#self.concepto_debe.setEnabled(False) and #self.concepto_debe.setText("")
        
        validadcion = validaciones(date,deber,haberes)
        if validadcion != "Validación exitosa.":
            mensaje_ingreso_datos("Error de validación", "Verifique los datos por favor")
            return        
        try:
            db = conectar_base_de_datos()
            cursor = db.cursor()
            query = f"UPDATE contabilidad SET fecha = '{date}', concepto_debe = '{descripcion}', concepto_haber = '{descripcion_h}', debe = '{deber}', haber = '{haberes}' WHERE id_concepto = '{id_concepto}'"
            cursor.execute(query)
            db.commit()
            print(f"\nREGISTRO: {date}, {descripcion},{descripcion_h}, {deber}, {haberes}")
            
            if cursor:
                ingreso_datos("Registro de Ingresos-Egresos","Datos cargados correctamente")
                limpiarCampos(self,QDate)
                self.tablaGastos.clearSelection() # Deselecciona la fila
            else:
                ingreso_datos("Registro de Ingresos-Egresos","Datos no cargados correctamente")
                limpiarCampos(self,QDate)
                self.concepto_debe.setEnabled(True)
                self.concepto_haber.setEnabled(True)
                self.tablaGastos.clearSelection() # Deselecciona la fila
            cursor.close()
            db.close()
            
        except Error as ex:
            errorConsulta("Registro de Ingresos-Egresos",f"Error en la consulta: {str(ex)}")
            print("Error executing the query", ex)
    
    def visualizacion_datos(self):
        principio = self.fecha_periodo.date().toString("yyyy-MM-dd")
        if not principio or principio > QDate.currentDate().toString("yyyy-MM-dd"):
            self.fecha_periodo.setDate(QDate.currentDate())
            mensaje_ingreso_datos("Registro de Ingresos-Egresos","La fecha de inicio debe ser la actual o una fecha anterior.")
            return 
        
        final = self.fecha_fin.date().toString("yyyy-MM-dd")
        if final > QDate.currentDate().toString("yyyy-MM-dd") or final < principio:
            self.fecha_fin.setDate(QDate.currentDate())
            mensaje_ingreso_datos("Registro de Ingresos-Egresos","La fecha de inicio debe ser la actual o una fecha anterior.\nLa fecha final de periodo debe ser igual o no anterior a la fecha de inicio de periodo")
            return
        
        try:
            db = conectar_base_de_datos()
            cursor = db.cursor()
            
            cursor.execute(f"SELECT id_concepto AS REGISTRO, fecha, concepto_debe, concepto_haber, debe AS 'DEBE ($)', haber AS 'HABER ($)' FROM contabilidad WHERE fecha BETWEEN '{principio}' AND '{final}' AND habilitado = 1 ORDER BY fecha ASC")
            busqueda = cursor.fetchall()
            if len(busqueda) > 0:
                ingreso_datos("Registro de Ingresos-Egresos",f"Se encontraron {len(busqueda)} coincidencias.")
                self.fecha_periodo.setDate(QDate.currentDate())
                self.fecha_fin.setDate(QDate.currentDate())
                
                tabla_contabilidad(self,cursor,busqueda,QHeaderView,QTableWidget,QAbstractItemView,QTableWidgetItem,QDate,Qt)
            else:
                ingreso_datos("Registro de Ingresos-Egresos",f"Se encontraron {len(busqueda)} coincidencias.")
                self.fecha_periodo.setDate(QDate.currentDate())
                self.fecha_fin.setDate(QDate.currentDate())
            
            cursor.close()
            db.close()
        except Error as ex:
            errorConsulta("Registro de Ingresos-Egresos",f"Error en la consulta: {str(ex)}")
            print("Error executing the query", ex)
           
    def eliminar_datos(self):#-----------------
        if not self.tablaGastos.currentItem():
            mensaje_ingreso_datos("Registro de Ingresos-Egresos","Debe seleccionar el registro que desea eliminar")
            return
        
        registro = self.tablaGastos.currentItem().row()
        idconcepto = int(self.tablaGastos.item(registro,0).text())

        # Verifica si ya se presionó el botón sin seleccionar una nueva fila
        if self.eliminacion_pendiente:
            mensaje_ingreso_datos("Registro de Ingresos-Egresos", "Debe seleccionar una nueva fila antes de eliminar otro registro")
            return
        
        elimContabilidad = inicio("Registro de Ingresos-Egresos","¿Desea eliminar el registro?")
        if elimContabilidad == QMessageBox.StandardButton.Yes:
            elimContabilidad2 = inicio("Registro de Ingresos-Egresos","¿Seguro que desea eliminar el registro")
            if elimContabilidad2 == QMessageBox.StandardButton.Yes:
                try:
                    db = conectar_base_de_datos()
                    cursor = db.cursor()
                    cursor.execute("UPDATE contabilidad SET habilitado = 0 WHERE id_concepto = %s", (idconcepto,))
                    # cursor.execute(f"DELETE FROM contabilidad WHERE id_concepto = {idconcepto}")
                    db.commit()
                    
                    if cursor.rowcount > 0:
                        ingreso_datos("Registro de Ingresos-Egresos","Registro eliminado")
                        self.tablaGastos.clearSelection()  # Deseleccionar la fila eliminada
                        self.tablaGastos.removeRow(registro)
                        
                        if self.tablaGastos.rowCount() == 1:
                            self.tablaGastos.setRowCount(0)  # Eliminar el registro de las sumatorias

                        limpiarCampos(self,QDate)                    
                    else:
                        ingreso_datos("Registro de Ingresos-Egresos","Registro no eliminado")
                                
                    cursor.close()
                    db.close()
                    self.eliminacion_pendiente = True  # Marca que se ha eliminado un registro y no se ha seleccionado una nueva fila

                except Error as ex:
                    errorConsulta("Registro de Ingresos-Egresos",f"Error en la consulta: {str(ex)}")
                    print("Error executing the query", ex)
            else:
                print("NO se elimino")
        else:
            print("NO se eliimina contabilidad")
            
    def limp_tabla(self):
        clear_tabla(self)
      
    def tabla_resumen(self):
        tabla_libroDiario_CONTABILIDAD(self,Workbook,Font,PatternFill,Border,Side,numbers,QFileDialog)
    
    def seleccionar_fila(self):
        # Método para manejar la selección de una nueva fila
        self.eliminacion_pendiente = False  # Restablece el estado al seleccionar una nueva fila
